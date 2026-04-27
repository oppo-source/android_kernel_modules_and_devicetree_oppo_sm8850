#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
进程内存详细分析工具
- 通过进程名或PID分析进程的详细内存占用
- 重点关注独享内存和异常内存占用
- 生成CSV和Excel文件便于分享和分析
- 支持从dumpsys_pss.csv批量读取进程列表
"""

import argparse
import csv
import datetime
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple


# ============================================================================
# 基础工具函数（整合自 process_memory_detail.py）
# ============================================================================

def sanitize_filename(name: str) -> str:
    """
    清理文件名，移除Windows不允许的字符

    Windows不允许的字符: < > : " / \ | ? *
    """
    # 替换Windows不允许的字符为下划线
    invalid_chars = '<>:"/\\|?*'
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    # 移除首尾空格和点
    sanitized = sanitized.strip(' .')
    # 如果为空，使用默认名称
    if not sanitized:
        sanitized = "process"
    return sanitized


def get_tools_path(relative_path: str) -> str:
    """获取工具输出路径"""
    base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    os.makedirs(full_path, exist_ok=True)
    return full_path


def _run(cmd: List[str], check: bool = False, timeout: Optional[int] = None) -> subprocess.CompletedProcess:
    """执行命令"""
    return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, check=check)


def _adb_base(serial: Optional[str]) -> List[str]:
    """构建adb基础命令"""
    base = ["adb"]
    if serial:
        base += ["-s", serial]
    return base


def adb_shell(
    shell_cmd: str,
    serial: Optional[str] = None,
    use_su: bool = False,
    timeout: Optional[int] = None,
) -> subprocess.CompletedProcess:
    """
    执行: adb shell <cmd>
    如果 use_su=True: adb shell su -c "<cmd>"
    """
    base = _adb_base(serial) + ["shell"]
    if use_su:
        return _run(base + ["su", "-c", shell_cmd], timeout=timeout)
    return _run(base + [shell_cmd], timeout=timeout)


def adb_get_state(serial: Optional[str]) -> bool:
    """检查adb设备状态"""
    try:
        r = _run(_adb_base(serial) + ["get-state"])
        return r.returncode == 0 and (r.stdout or "").strip() in ("device", "recovery", "sideload")
    except Exception:
        return False


def parse_pidof_output(s: str) -> List[int]:
    """解析pidof输出"""
    s = (s or "").strip()
    if not s:
        return []
    out: List[int] = []
    for tok in s.split():
        tok = tok.strip()
        if tok.isdigit():
            out.append(int(tok))
    return out


def find_pids(process_name: str, serial: Optional[str], use_su: bool) -> List[int]:
    """通过进程名查找PID列表"""
    # 先尝试pidof
    r = adb_shell(f"pidof {process_name}", serial=serial, use_su=use_su)
    pids = parse_pidof_output(r.stdout)
    if pids:
        return pids

    # 回退到ps
    r2 = adb_shell("ps -A", serial=serial, use_su=use_su)
    lines = (r2.stdout or "").splitlines()
    if not lines:
        return []

    header = lines[0]
    pid_idx = None
    name_idx = None
    cols = header.split()
    for i, c in enumerate(cols):
        if c.upper() == "PID":
            pid_idx = i
        if c.upper() in ("NAME", "CMDLINE", "COMMAND"):
            name_idx = i

    matched: List[int] = []
    for ln in lines[1:]:
        parts = ln.split()
        if pid_idx is not None and len(parts) > pid_idx:
            pid_str = parts[pid_idx]
        else:
            m = re.search(r"\b(\d+)\b", ln)
            pid_str = m.group(1) if m else ""

        name = ""
        if name_idx is not None and len(parts) > name_idx:
            name = parts[name_idx]
        elif parts:
            name = parts[-1]

        if name == process_name and pid_str.isdigit():
            matched.append(int(pid_str))

    return matched


def safe_write_text(path: str, content: str) -> None:
    """安全写入文本文件"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(content or "")


def kb_to_mb(kb: int) -> float:
    """KB转MB"""
    return kb / 1024.0


# ============================================================================
# 内存映射分类和解析（整合自 process_memory_detail.py）
# ============================================================================

def classify_mapping(pathname: str, bracket_name: str) -> str:
    """
    将VMA分类到粗粒度类别，帮助回答：
    "内存分别分配到什么地方 / 各占多少"
    """
    p = (pathname or "").strip()
    b = (bracket_name or "").strip()

    # Bracket区域
    if b:
        if b == "[heap]":
            return "anon_heap"
        if b.startswith("[anon:libc_malloc"):
            return "native_heap_malloc"
        if b.startswith("[anon:linker_alloc"):
            return "linker_alloc"
        if b.startswith("[anon:.bss"):
            return "anon_bss"
        if b.startswith("[anon:bionic_alloc"):
            return "bionic_alloc"
        if b.startswith("[anon:stack_and_tls:"):
            return "stack_tls"
        if b.startswith("[stack"):
            return "stack"
        if b in ("[vdso]", "[vvar]", "[vsyscall]"):
            return "kernel_vdso"
        if b == "[anon:thread signal stack]":
            return "thread_signal_stack"
        return f"bracket:{b}"

    if not p:
        return "anonymous"

    # 设备映射
    if p.startswith("/dev/"):
        if "ashmem" in p:
            return "ashmem"
        if "dmabuf" in p:
            return "dmabuf"
        if "/dev/ion" in p or "ion" in p:
            return "ion"
        if "kgsl" in p:
            return "gpu_kgsl"
        if "graphics" in p:
            return "gfx_dev"
        return "dev_other"

    # APK/JAR/DEX/OAT/ART/VDEX
    for ext, cat in [
        (".apk", "apk_mmap"),
        (".jar", "jar_mmap"),
        (".dex", "dex_mmap"),
        (".oat", "oat_mmap"),
        (".art", "art_mmap"),
        (".vdex", "vdex_mmap"),
    ]:
        if p.endswith(ext):
            return cat

    # 共享库
    if p.endswith(".so"):
        if "/apex/" in p:
            return "so_mmap_apex"
        if p.startswith("/system/"):
            return "so_mmap_system"
        if p.startswith("/vendor/"):
            return "so_mmap_vendor"
        return "so_mmap_other"

    # 字体/资源
    if p.endswith(".ttf") or p.endswith(".otf"):
        return "font_mmap"

    # 系统分区
    if p.startswith("/system/"):
        return "file_system"
    if p.startswith("/vendor/"):
        return "file_vendor"
    if p.startswith("/apex/"):
        return "file_apex"
    if p.startswith("/data/"):
        return "file_data"

    return "file_other"


@dataclass
class SmapsMetrics:
    """smaps指标"""
    size_kb: int = 0
    rss_kb: int = 0
    pss_kb: int = 0
    shared_clean_kb: int = 0
    shared_dirty_kb: int = 0
    private_clean_kb: int = 0
    private_dirty_kb: int = 0
    swap_kb: int = 0
    swap_pss_kb: int = 0

    def add(self, other: "SmapsMetrics") -> None:
        """累加指标"""
        self.size_kb += other.size_kb
        self.rss_kb += other.rss_kb
        self.pss_kb += other.pss_kb
        self.shared_clean_kb += other.shared_clean_kb
        self.shared_dirty_kb += other.shared_dirty_kb
        self.private_clean_kb += other.private_clean_kb
        self.private_dirty_kb += other.private_dirty_kb
        self.swap_kb += other.swap_kb
        self.swap_pss_kb += other.swap_pss_kb


@dataclass
class SmapsEntry:
    """smaps条目"""
    start: str
    end: str
    perms: str
    offset: str
    dev: str
    inode: str
    pathname: str
    bracket_name: str
    metrics: SmapsMetrics

    @property
    def key_path(self) -> str:
        return self.pathname or self.bracket_name or ""


SMAPS_HEADER_RE = re.compile(
    r"^([0-9a-fA-F]+)-([0-9a-fA-F]+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s*(.*)$"
)


def parse_smaps(text: str) -> List[SmapsEntry]:
    """解析smaps文本"""
    entries: List[SmapsEntry] = []
    cur: Optional[SmapsEntry] = None

    def _commit() -> None:
        nonlocal cur
        if cur is not None:
            entries.append(cur)
            cur = None

    for raw in (text or "").splitlines():
        line = raw.rstrip("\n")
        m = SMAPS_HEADER_RE.match(line)
        if m:
            _commit()
            start, end, perms, offset, dev, inode, path = m.groups()
            path = (path or "").strip()
            bracket = ""
            if path.startswith("[") and path.endswith("]"):
                bracket = path
                path = ""
            cur = SmapsEntry(
                start=start,
                end=end,
                perms=perms,
                offset=offset,
                dev=dev,
                inode=inode,
                pathname=path,
                bracket_name=bracket,
                metrics=SmapsMetrics(),
            )
            continue

        if cur is None:
            continue

        # 解析指标行: "Pss:                123 kB"
        m2 = re.match(r"^(\w+):\s+(\d+)\s+kB\s*$", line)
        if not m2:
            continue
        k, v = m2.group(1), int(m2.group(2))
        if k == "Size":
            cur.metrics.size_kb = v
        elif k == "Rss":
            cur.metrics.rss_kb = v
        elif k == "Pss":
            cur.metrics.pss_kb = v
        elif k == "Shared_Clean":
            cur.metrics.shared_clean_kb = v
        elif k == "Shared_Dirty":
            cur.metrics.shared_dirty_kb = v
        elif k == "Private_Clean":
            cur.metrics.private_clean_kb = v
        elif k == "Private_Dirty":
            cur.metrics.private_dirty_kb = v
        elif k == "Swap":
            cur.metrics.swap_kb = v
        elif k == "SwapPss":
            cur.metrics.swap_pss_kb = v

    _commit()
    return entries


def parse_smaps_rollup(text: str) -> Dict[str, int]:
    """解析smaps_rollup文本"""
    out: Dict[str, int] = {}
    for raw in (text or "").splitlines():
        line = raw.strip()
        m = re.match(r"^(\w+):\s+(\d+)\s+kB\s*$", line)
        if not m:
            continue
        out[m.group(1)] = int(m.group(2))
    return out


# ============================================================================
# 进程解析和批量读取
# ============================================================================

def resolve_process(process_name: Optional[str], pid: Optional[int], serial: Optional[str]) -> Tuple[List[int], str]:
    """
    解析进程名或PID，返回PID列表和进程名

    Returns:
        (pid_list, process_name)
    """
    if pid is not None:
        # 通过PID获取进程名
        r = adb_shell(f"ps -p {pid}", serial=serial, use_su=False, timeout=10)
        if r.returncode == 0:
            lines = (r.stdout or "").strip().splitlines()
            if len(lines) >= 2:
                parts = lines[1].split()
                if parts:
                    name = parts[-1]
                    return [pid], name
        return [pid], f"process_{pid}"

    if process_name:
        pids = find_pids(process_name, serial=serial, use_su=False)
        if pids:
            return pids, process_name

    return [], ""


def get_process_name_from_pid(pid: int, serial: Optional[str]) -> str:
    """通过PID获取进程名"""
    r = adb_shell(f"ps -p {pid}", serial=serial, use_su=False, timeout=10)
    if r.returncode == 0:
        lines = (r.stdout or "").strip().splitlines()
        if len(lines) >= 2:
            parts = lines[1].split()
            if parts:
                return parts[-1]
    return f"process_{pid}"


def read_pids_from_dumpsys_pss(csv_file: str) -> List[Tuple[int, str]]:
    """
    从dumpsys_pss.csv读取PID列表

    Returns:
        List of (pid, process_name) tuples
    """
    pids = []
    if not os.path.isfile(csv_file):
        print(f"WARNING: 文件不存在: {csv_file}")
        return pids

    try:
        with open(csv_file, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                pid_str = row.get("pid", "").strip()
                process_name = row.get("process_name", "").strip()

                if pid_str and pid_str.isdigit():
                    pid = int(pid_str)
                    # 跳过TOTAL行
                    if process_name.upper() != "TOTAL":
                        pids.append((pid, process_name))

        print(f"从 {csv_file} 读取到 {len(pids)} 个进程")
    except Exception as e:
        print(f"读取 {csv_file} 时出错: {e}")

    return pids


# ============================================================================
# 内存分析核心函数
# ============================================================================

def calculate_exclusive_memory(metrics: SmapsMetrics) -> int:
    """计算独享内存 = Private_Dirty + Private_Clean"""
    return metrics.private_dirty_kb + metrics.private_clean_kb


def detect_anomalies(
    entry: SmapsEntry,
    total_pss_kb: int,
    category_pss_kb: Dict[str, int],
    category_exclusive_kb: Dict[str, int],
) -> Tuple[bool, str]:
    """
    检测异常内存占用

    Returns:
        (is_anomaly, anomaly_type)
    """
    if total_pss_kb == 0:
        return False, ""

    exclusive = calculate_exclusive_memory(entry.metrics)
    category = classify_mapping(entry.pathname, entry.bracket_name)

    anomalies = []

    # 规则1: 单个映射独享内存 > PSS的10%
    if exclusive > 0 and exclusive > total_pss_kb * 0.1:
        anomalies.append("独享内存异常大")

    # 规则2: 单个映射PSS > PSS的5%
    if entry.metrics.pss_kb > total_pss_kb * 0.05:
        anomalies.append("PSS异常大")

    # 规则3: 类别独享内存 > PSS的30%
    cat_exclusive = category_exclusive_kb.get(category, 0)
    if cat_exclusive > total_pss_kb * 0.3:
        anomalies.append("类别独享内存异常大")

    if anomalies:
        return True, "; ".join(anomalies)

    return False, ""


def verify_pss(
    smaps_rollup_pss: Optional[int],
    category_pss_sum: int,
    detailed_pss_sum: int,
) -> List[Dict[str, any]]:
    """
    PSS验证

    Returns:
        验证结果列表
    """
    results = []

    # 验证1: smaps_rollup vs 详细映射总和
    if smaps_rollup_pss is not None:
        diff = detailed_pss_sum - smaps_rollup_pss
        diff_pct = (diff / smaps_rollup_pss * 100.0) if smaps_rollup_pss > 0 else 0.0

        if abs(diff_pct) <= 5.0:
            status = "通过"
        elif abs(diff_pct) <= 10.0:
            status = "警告"
        else:
            status = "失败"

        results.append({
            "验证项": "smaps_rollup_PSS",
            "基准值(kB)": smaps_rollup_pss,
            "计算值(kB)": detailed_pss_sum,
            "差异(kB)": diff,
            "差异百分比(%)": f"{diff_pct:.4f}",
            "验证结果": status,
        })

    # 验证2: 类别汇总 vs 详细映射总和
    diff2 = category_pss_sum - detailed_pss_sum
    diff_pct2 = (diff2 / detailed_pss_sum * 100.0) if detailed_pss_sum > 0 else 0.0

    if abs(diff_pct2) <= 1.0:
        status2 = "通过"
    else:
        status2 = "警告"

    results.append({
        "验证项": "类别汇总PSS",
        "基准值(kB)": category_pss_sum,
        "计算值(kB)": detailed_pss_sum,
        "差异(kB)": diff2,
        "差异百分比(%)": f"{diff_pct2:.4f}",
        "验证结果": status2,
    })

    return results


# ============================================================================
# CSV文件生成
# ============================================================================

def write_summary_csv(
    output_path: str,
    process_name: str,
    pid: int,
    now_str: str,
    smaps_rollup: Dict[str, int],
    category_totals: Dict[str, SmapsMetrics],
    verification_results: List[Dict[str, any]],
    entries: Optional[List[SmapsEntry]] = None,
) -> None:
    """写入汇总CSV文件（包含总览+验证+类别汇总）"""
    rows = []

    # 部分1: 进程总览
    rows.append(["# 进程内存分析总览"])
    rows.append([
        "进程名", "进程ID", "采样时间",
        "PSS总计(kB)", "PSS总计(MB)",
        "RSS总计(kB)", "RSS总计(MB)",
        "独享内存总计(kB)", "独享内存总计(MB)", "独享内存占比(%)",
        "共享内存总计(kB)", "共享内存总计(MB)"
    ])

    pss_total = smaps_rollup.get("Pss", 0)
    rss_total = smaps_rollup.get("Rss", 0)
    exclusive_total = smaps_rollup.get("Private_Dirty", 0) + smaps_rollup.get("Private_Clean", 0)
    shared_total = smaps_rollup.get("Shared_Dirty", 0) + smaps_rollup.get("Shared_Clean", 0)
    exclusive_pct = (exclusive_total / pss_total * 100.0) if pss_total > 0 else 0.0

    rows.append([
        process_name, pid, now_str,
        pss_total, f"{kb_to_mb(pss_total):.2f}",
        rss_total, f"{kb_to_mb(rss_total):.2f}",
        exclusive_total, f"{kb_to_mb(exclusive_total):.2f}", f"{exclusive_pct:.2f}",
        shared_total, f"{kb_to_mb(shared_total):.2f}"
    ])
    rows.append([])  # 空行分隔

    # 部分2: PSS验证
    if verification_results:
        rows.append(["# PSS验证"])
        rows.append([
            "验证项", "基准值(kB)", "计算值(kB)", "差异(kB)", "差异百分比(%)", "验证结果"
        ])
        for v in verification_results:
            rows.append([
                v["验证项"], v["基准值(kB)"], v["计算值(kB)"],
                v["差异(kB)"], v["差异百分比(%)"], v["验证结果"]
            ])
        rows.append([])  # 空行分隔

    # 部分3: 按类别汇总（按独享内存排序）
    rows.append(["# 按类别汇总（按独享内存排序）"])
    rows.append([
        "类别名称", "PSS(kB)", "PSS(MB)", "PSS占比(%)",
        "独享内存(kB)", "独享内存(MB)", "独享内存占比(%)",
        "共享内存(kB)", "共享内存(MB)",
        "映射数量", "异常标记"
    ])

    # 计算总PSS用于计算占比
    total_pss = smaps_rollup.get("Pss", 0)
    total_exclusive = smaps_rollup.get("Private_Dirty", 0) + smaps_rollup.get("Private_Clean", 0)

    # 统计每个类别的映射数量
    category_counts: Dict[str, int] = {}
    if entries:
        for entry in entries:
            category = classify_mapping(entry.pathname, entry.bracket_name)
            category_counts[category] = category_counts.get(category, 0) + 1

    # 按独享内存排序
    sorted_categories = sorted(
        category_totals.items(),
        key=lambda kv: calculate_exclusive_memory(kv[1]),
        reverse=True
    )

    for cat, metrics in sorted_categories:
        exclusive = calculate_exclusive_memory(metrics)
        shared = metrics.shared_dirty_kb + metrics.shared_clean_kb
        pss_pct = (metrics.pss_kb / total_pss * 100.0) if total_pss > 0 else 0.0
        exclusive_pct = (exclusive / total_exclusive * 100.0) if total_exclusive > 0 else 0.0

        # 异常标记：类别独享内存 > PSS的30%
        is_anomaly = exclusive > total_pss * 0.3 if total_pss > 0 else False
        anomaly_mark = "异常" if is_anomaly else "正常"

        mapping_count = category_counts.get(cat, 0)

        rows.append([
            cat,
            metrics.pss_kb, f"{kb_to_mb(metrics.pss_kb):.2f}", f"{pss_pct:.2f}",
            exclusive, f"{kb_to_mb(exclusive):.2f}", f"{exclusive_pct:.2f}",
            shared, f"{kb_to_mb(shared):.2f}",
            mapping_count,
            anomaly_mark
        ])

    # 写入CSV
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            writer.writerow(row)

    print(f"  ✓ 汇总文件已保存: {output_path}")


def write_detailed_csv(
    output_path: str,
    entries: List[SmapsEntry],
    total_pss_kb: int,
    category_totals: Dict[str, SmapsMetrics],
) -> None:
    """写入详细映射CSV文件"""
    rows = []

    # 表头
    rows.append([
        "地址范围", "路径/名称", "类别", "权限",
        "PSS(kB)", "RSS(kB)",
        "Private_Dirty(kB)", "Private_Clean(kB)",
        "Shared_Dirty(kB)", "Shared_Clean(kB)",
        "独享内存(kB)", "共享进程数",
        "是否独享内存", "是否异常", "异常类型"
    ])

    # 计算类别汇总（用于异常检测）
    category_pss = {cat: m.pss_kb for cat, m in category_totals.items()}
    category_exclusive = {
        cat: calculate_exclusive_memory(m) for cat, m in category_totals.items()
    }

    # 按独享内存排序
    sorted_entries = sorted(
        entries,
        key=lambda e: calculate_exclusive_memory(e.metrics),
        reverse=True
    )

    for entry in sorted_entries:
        exclusive = calculate_exclusive_memory(entry.metrics)
        category = classify_mapping(entry.pathname, entry.bracket_name)
        is_exclusive = "是" if exclusive > 0 else "否"

        is_anomaly, anomaly_type = detect_anomalies(
            entry, total_pss_kb, category_pss, category_exclusive
        )
        anomaly_mark = "是" if is_anomaly else "否"

        path_display = entry.pathname or entry.bracket_name or ""
        address_range = f"{entry.start}-{entry.end}"

        rows.append([
            address_range,
            path_display,
            category,
            entry.perms,
            entry.metrics.pss_kb,
            entry.metrics.rss_kb,
            entry.metrics.private_dirty_kb,
            entry.metrics.private_clean_kb,
            entry.metrics.shared_dirty_kb,
            entry.metrics.shared_clean_kb,
            exclusive,
            "",  # 共享进程数（暂不实现）
            is_exclusive,
            anomaly_mark,
            anomaly_type
        ])

    # 写入CSV
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            writer.writerow(row)

    print(f"  ✓ 详细文件已保存: {output_path}")


# ============================================================================
# Excel转换和美化
# ============================================================================

def csv_to_excel(csv_filename: str, excel_filename: Optional[str] = None) -> None:
    """将CSV转换为Excel，支持多部分CSV（用注释行分隔）"""
    try:
        import pandas as pd
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  ⚠ WARNING: pandas/openpyxl未安装，跳过Excel转换: {csv_filename}")
        return

    try:
        if not os.path.isfile(csv_filename):
            print(f"  ⚠ WARNING: CSV文件不存在: {csv_filename}")
            return

        if excel_filename is None:
            base_name = os.path.splitext(csv_filename)[0]
            excel_filename = f"{base_name}.xlsx"

        # 读取CSV
        rows = []
        with open(csv_filename, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)

        if not rows:
            print(f"  ⚠ WARNING: CSV文件为空: {csv_filename}")
            return

        # 创建Excel工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # 解析多部分CSV
        current_section = None
        current_header = None
        current_data = []
        section_count = 0

        for row in rows:
            if not row:
                # 空行：如果当前有数据，保存当前section
                if current_header and current_data:
                    section_count += 1
                    sheet_name = current_section.replace("# ", "").replace(" ", "_")[:31] if current_section else f"Sheet{section_count}"
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(current_header)
                    for data_row in current_data:
                        ws.append(data_row)
                    current_data = []
                    current_header = None
                    current_section = None
                continue

            # 检查是否是注释行（section标题）
            if row[0].startswith("#"):
                # 保存上一个section
                if current_header and current_data:
                    section_count += 1
                    sheet_name = current_section.replace("# ", "").replace(" ", "_")[:31] if current_section else f"Sheet{section_count}"
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(current_header)
                    for data_row in current_data:
                        ws.append(data_row)

                # 开始新section
                current_section = row[0]
                current_header = None
                current_data = []
                continue

            # 数据行
            if current_header is None:
                # 这是表头
                current_header = row
            else:
                # 这是数据行
                current_data.append(row)

        # 保存最后一个section
        if current_header and current_data:
            section_count += 1
            sheet_name = current_section.replace("# ", "").replace(" ", "_")[:31] if current_section else f"Sheet{section_count}"
            ws = wb.create_sheet(title=sheet_name)
            ws.append(current_header)
            for data_row in current_data:
                ws.append(data_row)

        # 如果没有找到任何section，使用简单模式
        if section_count == 0:
            ws = wb.create_sheet(title="数据")
            # 找到第一个非注释行作为表头
            header_idx = 0
            for i, row in enumerate(rows):
                if row and not (row[0].startswith("#") if row[0] else False):
                    header_idx = i
                    break
            header = rows[header_idx]
            data_rows = rows[header_idx + 1:]
            ws.append(header)
            for row in data_rows:
                if row and not (row[0].startswith("#") if row[0] else False):
                    ws.append(row)

        wb.save(excel_filename)
        print(f"  ✓ Excel文件已保存: {excel_filename}")

        # 美化Excel
        beautify_excel_file(excel_filename)

    except Exception as e:
        print(f"  ✗ 转换Excel时出错: {e}")


def beautify_excel_file(excel_filename: str, max_width: int = 50, min_width: int = 10, sample_rows: int = 2000) -> None:
    """美化Excel文件"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    try:
        if not os.path.isfile(excel_filename) or os.path.getsize(excel_filename) <= 0:
            return

        wb = load_workbook(excel_filename)

        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for ws in wb.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row <= 0 or max_col <= 0:
                continue

            # 冻结首行
            if max_row >= 2:
                ws.freeze_panes = "A2"

            # 表头样式
            ws.row_dimensions[1].height = 18
            for c in range(1, max_col + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动筛选
            try:
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"
            except Exception:
                pass

            # 数字列右对齐
            num_align = Alignment(horizontal="right", vertical="center")
            txt_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

            # 判断数字列
            numeric_cols = set()
            scan_rows = min(max_row, 1 + sample_rows)
            for c in range(1, max_col + 1):
                total_cnt = 0
                num_cnt = 0
                for r in range(2, scan_rows + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None or v == "":
                        continue
                    total_cnt += 1
                    if isinstance(v, (int, float)):
                        num_cnt += 1
                if total_cnt > 0 and (num_cnt / float(total_cnt)) >= 0.8:
                    numeric_cols.add(c)

            # 应用对齐
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if c in numeric_cols and isinstance(cell.value, (int, float)):
                        cell.alignment = num_align
                    else:
                        if cell.alignment is None or cell.alignment.horizontal is None:
                            cell.alignment = txt_align

            # 列宽自适应
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                best = 0
                for r in range(1, scan_rows + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    s = str(v)
                    best = max(best, len(s))
                width = max(min_width, min(max_width, best + 2))
                ws.column_dimensions[col_letter].width = width

        wb.save(excel_filename)
    except Exception as e:
        print(f"  ⚠ 美化Excel时出错: {e}")


# ============================================================================
# 主函数
# ============================================================================

def analyze_single_process(
    process_name: str,
    pid: int,
    serial: Optional[str],
    use_su: bool,
    base_out_dir: Optional[str],
    verbose: bool = True,
) -> bool:
    """
    分析单个进程的内存占用

    Returns:
        True if successful, False otherwise
    """
    if verbose:
        print(f"\n{'='*80}")
        print(f"开始分析进程: {process_name} (PID: {pid})")
        print(f"{'='*80}")

    now = datetime.datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")

    # 确定输出目录（清理进程名中的非法字符）
    sanitized_name = sanitize_filename(process_name)
    if base_out_dir:
        out_dir = base_out_dir
    else:
        out_dir = get_tools_path(os.path.join("out", "process_memory", f"{sanitized_name}_{pid}_mem"))

    os.makedirs(out_dir, exist_ok=True)
    raw_dir = os.path.join(out_dir, "raw")
    os.makedirs(raw_dir, exist_ok=True)

    if verbose:
        print(f"\n[步骤1] 准备输出目录")
        print(f"  输出目录: {out_dir}")
        print(f"  原始数据目录: {raw_dir}")

    # 1. 采集原始数据
    if verbose:
        print(f"\n[步骤2] 采集原始数据（这可能需要一些时间...）")
        print(f"  正在读取 /proc/{pid}/smaps_rollup...")

    smaps_rollup_path = os.path.join(raw_dir, "smaps_rollup.txt")
    r_roll = adb_shell(f"cat /proc/{pid}/smaps_rollup", serial=serial, use_su=use_su, timeout=60)
    safe_write_text(
        smaps_rollup_path,
        (r_roll.stdout or "") + (("\n\n[stderr]\n" + r_roll.stderr) if r_roll.stderr else "")
    )
    smaps_rollup = parse_smaps_rollup(r_roll.stdout or "")

    if not smaps_rollup:
        if verbose:
            print(f"  ✗ 无法解析 smaps_rollup，可能权限不足。尝试使用 --su")
        return False

    if verbose:
        pss_total = smaps_rollup.get("Pss", 0)
        print(f"  ✓ smaps_rollup读取成功")
        print(f"    进程PSS总计: {pss_total} kB ({kb_to_mb(pss_total):.2f} MB)")

    if verbose:
        print(f"  正在读取 /proc/{pid}/smaps（这可能需要较长时间，请耐心等待...）")

    smaps_path = os.path.join(raw_dir, "smaps.txt")
    r_smaps = adb_shell(f"cat /proc/{pid}/smaps", serial=serial, use_su=use_su, timeout=180)
    safe_write_text(
        smaps_path,
        (r_smaps.stdout or "") + (("\n\n[stderr]\n" + r_smaps.stderr) if r_smaps.stderr else "")
    )
    entries = parse_smaps(r_smaps.stdout or "")

    if not entries:
        if verbose:
            print(f"  ✗ 无法解析 smaps，可能权限不足或数据为空")
        return False

    if verbose:
        print(f"  ✓ smaps读取成功，共解析到 {len(entries)} 个内存映射")

    if verbose:
        print(f"  正在读取 dumpsys meminfo {pid}...")

    dumpsys_path = os.path.join(raw_dir, "dumpsys_meminfo.txt")
    r_dm = adb_shell(f"dumpsys meminfo {pid}", serial=serial, use_su=False, timeout=60)
    safe_write_text(
        dumpsys_path,
        (r_dm.stdout or "") + (("\n\n[stderr]\n" + r_dm.stderr) if r_dm.stderr else "")
    )

    if verbose:
        print(f"  ✓ dumpsys meminfo读取成功")

    if verbose:
        print(f"  正在读取 /proc/{pid}/status...")

    status_path = os.path.join(raw_dir, "status.txt")
    r_status = adb_shell(f"cat /proc/{pid}/status", serial=serial, use_su=use_su, timeout=30)
    safe_write_text(
        status_path,
        (r_status.stdout or "") + (("\n\n[stderr]\n" + r_status.stderr) if r_status.stderr else "")
    )

    if verbose:
        print(f"  ✓ status读取成功")
        print(f"  所有原始数据已保存到: {raw_dir}")

    # 2. 解析和分类
    if verbose:
        print(f"\n[步骤3] 解析和分类内存映射")
        print(f"  正在对 {len(entries)} 个内存映射进行分类...")

    category_totals: Dict[str, SmapsMetrics] = {}

    for entry in entries:
        category = classify_mapping(entry.pathname, entry.bracket_name)
        if category not in category_totals:
            category_totals[category] = SmapsMetrics()
        category_totals[category].add(entry.metrics)

    if verbose:
        print(f"  ✓ 分类完成，共 {len(category_totals)} 个类别")
        print(f"  主要类别:")
        sorted_cats = sorted(category_totals.items(), key=lambda kv: kv[1].pss_kb, reverse=True)[:5]
        for cat, metrics in sorted_cats:
            exclusive = calculate_exclusive_memory(metrics)
            print(f"    - {cat}: PSS={metrics.pss_kb}kB, 独享={exclusive}kB")

    # 计算汇总值
    total_pss_from_entries = sum(e.metrics.pss_kb for e in entries)
    total_pss_from_categories = sum(m.pss_kb for m in category_totals.values())
    smaps_rollup_pss = smaps_rollup.get("Pss")

    # 3. PSS验证
    if verbose:
        print(f"\n[步骤4] PSS验证（确保分析结果准确性）")

    verification_results = verify_pss(
        smaps_rollup_pss,
        total_pss_from_categories,
        total_pss_from_entries
    )

    if verbose:
        for v in verification_results:
            status_icon = "✓" if v["验证结果"] == "通过" else "⚠" if v["验证结果"] == "警告" else "✗"
            print(f"  {status_icon} {v['验证项']}: {v['验证结果']} (差异: {v['差异百分比(%)']}%)")

    # 4. 生成CSV文件
    if verbose:
        print(f"\n[步骤5] 生成CSV文件")

    summary_csv = os.path.join(out_dir, "process_memory_summary.csv")
    detailed_csv = os.path.join(out_dir, "process_memory_detailed.csv")

    write_summary_csv(
        summary_csv,
        process_name,
        pid,
        now_str,
        smaps_rollup,
        category_totals,
        verification_results,
        entries
    )

    write_detailed_csv(
        detailed_csv,
        entries,
        smaps_rollup_pss or total_pss_from_entries,
        category_totals
    )

    # 5. 转换为Excel
    if verbose:
        print(f"\n[步骤6] 转换为Excel文件")

    csv_to_excel(summary_csv)
    csv_to_excel(detailed_csv)

    if verbose:
        print(f"\n{'='*80}")
        print(f"分析完成！结果保存在: {out_dir}")
        print(f"{'='*80}")
        print(f"\n生成的文件:")
        print(f"  - process_memory_summary.csv/xlsx: 总览、验证、类别汇总")
        print(f"  - process_memory_detailed.csv/xlsx: 详细映射列表")
        print(f"  - raw/: 原始数据文件")
        print(f"\n分析建议:")
        print(f"  1. 先查看 summary 文件，了解总体情况和主要类别")
        print(f"  2. 重点关注'独享内存'列，这是可以优化的内存")
        print(f"  3. 查看 detailed 文件，筛选'是否独享内存=是'，按独享内存排序")
        print(f"  4. 查看'异常标记'列，关注异常大的内存占用")
        print(f"{'='*80}\n")

    return True


def main() -> int:
    parser = argparse.ArgumentParser(
        description="分析进程详细内存占用（重点关注独享内存和异常占用）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 分析单个进程（通过进程名）
  python process_memory_analysis.py -p camerahalserver

  # 分析单个进程（通过PID）
  python process_memory_analysis.py --pid 1234

  # 从dumpsys_pss.csv批量分析（不指定-p或--pid时自动使用）
  python process_memory_analysis.py

  # 使用su权限（如果需要）
  python process_memory_analysis.py -p system_server --su

  # 指定输出目录
  python process_memory_analysis.py -p system_server -o ./my_output
        """
    )
    parser.add_argument("-p", "--process", type=str, default=None, help="进程名，如: camerahalserver")
    parser.add_argument("--pid", type=int, default=None, help="进程ID")
    parser.add_argument("-o", "--out", type=str, default=None, help="输出目录（默认: tools/out/process_memory/）")
    parser.add_argument("--serial", type=str, default=None, help="adb设备序列号")
    parser.add_argument("--su", action="store_true", help="使用su权限读取/proc/<pid>/smaps*")
    parser.add_argument("--dumpsys-pss", type=str, default=None, help="dumpsys_pss.csv文件路径（默认: out/dumpsys_pss.csv）")
    parser.add_argument("--quiet", action="store_true", help="减少输出信息")
    args = parser.parse_args()

    # 检查adb连接
    if not adb_get_state(args.serial):
        print("ERROR: adb设备未就绪。请运行 'adb devices' 确保设备已连接。")
        return 2

    verbose = not args.quiet

    # 确定要分析的进程列表
    processes_to_analyze: List[Tuple[int, str]] = []

    if args.process or args.pid is not None:
        # 单进程模式
        if args.pid is not None:
            process_name = get_process_name_from_pid(args.pid, args.serial)
            processes_to_analyze = [(args.pid, process_name)]
        else:
            pids, process_name = resolve_process(args.process, None, args.serial)
            if not pids:
                print(f"ERROR: 无法找到进程 '{args.process}'")
                return 3
            processes_to_analyze = [(pid, process_name) for pid in pids]
    else:
        # 批量模式：从dumpsys_pss.csv读取
        if verbose:
            print("未指定进程名或PID，尝试从dumpsys_pss.csv读取进程列表...")

        dumpsys_pss_file = args.dumpsys_pss
        if not dumpsys_pss_file:
            # 默认路径
            default_path = get_tools_path(os.path.join("out", "dumpsys_pss.csv"))
            dumpsys_pss_file = default_path

        processes_to_analyze = read_pids_from_dumpsys_pss(dumpsys_pss_file)

        if not processes_to_analyze:
            print("ERROR: 未找到要分析的进程。请:")
            print("  1. 指定进程名: -p <process_name>")
            print("  2. 指定PID: --pid <pid>")
            print("  3. 或确保 dumpsys_pss.csv 文件存在且包含有效的pid列")
            return 1

    total_count = len(processes_to_analyze)
    if verbose:
        print(f"\n准备分析 {total_count} 个进程")
        if total_count > 1:
            print(f"进程列表:")
            for idx, (pid, pname) in enumerate(processes_to_analyze, 1):
                print(f"  {idx}. {pname} (PID: {pid})")

    # 分析每个进程
    success_count = 0
    for idx, (pid, process_name) in enumerate(processes_to_analyze, 1):
        if total_count > 1:
            print(f"\n{'='*80}")
            print(f"[进度: {idx}/{total_count}] 正在分析: {process_name} (PID: {pid})")
            print(f"{'='*80}")

        if analyze_single_process(
            process_name,
            pid,
            args.serial,
            args.su,
            args.out,
            verbose
        ):
            success_count += 1
            if total_count > 1:
                print(f"\n[进度: {idx}/{total_count}] ✓ {process_name} (PID: {pid}) 分析完成")
        else:
            if total_count > 1:
                print(f"\n[进度: {idx}/{total_count}] ✗ {process_name} (PID: {pid}) 分析失败")

    if verbose:
        print(f"\n{'='*80}")
        print(f"批量分析完成: 成功 {success_count}/{total_count} 个进程")
        print(f"{'='*80}")

    return 0 if success_count > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
