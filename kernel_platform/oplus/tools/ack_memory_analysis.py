import os
import re
import subprocess
from datetime import datetime
import argparse
import common
import csv
import pandas as pd
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def get_tools_path(relative_path):
    base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    if not os.path.exists(full_path):
        os.makedirs(full_path, exist_ok=True)
    return full_path


def parse_cmd_args():
    parser = argparse.ArgumentParser(description="module list parser")
    parser.add_argument('-l', '--list', type=str, help='modules list dir',default=get_tools_path("out/mtk_master"))
    parser.add_argument('-o', '--out', type=str, help='Kernel tmp out dir (default: tools/out)',
                        default=get_tools_path("out"))
    parser.add_argument('--dumpsys-pss-rows', type=int, default=None,
                        help='Number of rows to process for dumpsys_pss path lookup (default: all rows)')

    args = parser.parse_args()
    for key, value in vars(args).items():
        print(f"{key}: {value}")
    return args

def run_adb_lsmod(output_file):
    try:
        # 执行 adb shell lsmod -t 命令
        result = subprocess.run(['adb', 'shell', 'lsmod', '-t'], capture_output=True, text=True)

        # 解析命令输出
        lines = result.stdout.strip().split('\n')
        if not lines:
            print("命令输出为空")
            return

        # 提取表头
        header = lines[0].split()

        # 解析数据行
        parsed_data = []
        for line in lines[1:]:
            parts = line.split()
            # 确保每一行至少有 4 列，不足的部分用空字符串补齐
            parts += [''] * (4 - len(parts))

            module_name = parts[0]
            size = parts[1]
            used_by_count = parts[2]
            used_by_modules = parts[3]

            # 将 used_by_modules 用双引号括起来
            used_by_modules = f'"{used_by_modules}"'

            parsed_data.append([module_name, size, used_by_count, used_by_modules])

        # 按照字符进行排序
        sorted_data = sorted(parsed_data, key=lambda x: ''.join(x))

        # 将结果写入CSV文件
        with open(output_file, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)  # 写入表头
            writer.writerows(sorted_data)  # 写入排序后的数据

        print(f"命令执行成功，结果已保存到 {output_file}")
    except Exception as e:
        print(f"执行命令时出错: {e}")

def cat_proc_modules(output_file):
    try:
        # 执行 adb shell cat /proc/modules 命令
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/modules'], capture_output=True, text=True)

        # 解析命令输出
        lines = result.stdout.strip().split('\n')
        parsed_data = []

        for line in lines:
            parts = line.split()
            # 确保每一行至少有 7 列，不足的部分用空字符串补齐
            parts += [''] * (7 - len(parts))

            module_name = parts[0]
            size = parts[1]
            used_by_count = parts[2]
            used_by_modules = parts[3]
            status = parts[4]
            address = parts[5]
            flags = parts[6]

            # 将 used_by_modules 用双引号括起来
            used_by_modules = f'"{used_by_modules}"'

            parsed_data.append([module_name, size, used_by_count, used_by_modules, status, address, flags])

        # 按照第一列（Module Name）排序
        sorted_data = sorted(parsed_data, key=lambda x: x[0])

        # 将排序后的结果写入CSV文件
        with open(output_file, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            # 添加表头
            writer.writerow(['Module Name', 'Size', 'Used By Count', 'Used By Modules', 'Status', 'Address', 'Flags'])
            writer.writerows(sorted_data)

        print(f"命令执行成功，结果已保存到 {output_file}")
    except Exception as e:
        print(f"执行命令时出错: {e}")

def sort_csv_by_column(input_file, output_file, sort_column_index=0):
    try:
        # 读取CSV文件
        with open(input_file, 'r', newline='') as file:
            reader = csv.reader(file)
            header = next(reader)  # 读取表头
            data = list(reader)    # 读取剩余数据

        # 按照指定列排序
        sorted_data = sorted(data, key=lambda x: x[sort_column_index])

        # 将排序后的数据写入新的CSV文件
        with open(output_file, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)  # 写入表头
            writer.writerows(sorted_data)  # 写入排序后的数据

        print(f"排序成功，结果已保存到 {output_file}")
    except Exception as e:
        print(f"执行命令时出错: {e}")

def find_ko_files(directory, output_file):
    """查找目录中的所有 .ko 文件并写入输出文件，仅保留文件名并排序"""
    if not directory or not os.path.exists(directory):
        print(f"Warning: Directory not found or empty - {directory}, skipping .ko file search")
        # 创建空文件，避免后续处理出错
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        with open(output_file, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['file_name','module_name'])
        return

    print(f"Searching for .ko files in: {directory}")
    ko_files = []

    try:
        for root, _, files in os.walk(directory):
            print(f"Scanning: {root}")
            for file in files:
                if file.endswith('.ko'):
                    ko_files.append(file)  # 仅保留文件名
                    #print(f"Found: {file}")

        ko_files = list(set(ko_files))  # 去重
        ko_files.sort()  # 重新排序

        # 将结果写入CSV文件
        with open(output_file, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            # 添加表头
            writer.writerow(['file_name','module_name'])
            # 写入排序后的文件名
            for ko_file in ko_files:
                writer.writerow([ko_file])

        print(f"Results saved to: {output_file}")

    except Exception as e:
        print(f"Error processing files: {e}")

def process_ko_list(input_file, output_file):
    try:
        # 读取输入CSV文件
        with open(input_file, 'r', newline='') as file:
            reader = csv.DictReader(file)
            data = list(reader)

        # 处理每一行数据
        processed_data = []
        for row in data:
            file_name = row['file_name']
            # 去掉 .ko 后缀
            module_name = file_name.replace('.ko', '')
            # 将 - 替换为 _
            module_name = module_name.replace('-', '_')
            # 添加或更新 module_name 列
            row['module_name'] = module_name
            processed_data.append(row)

        # 获取表头
        fieldnames = reader.fieldnames
        if 'module_name' not in fieldnames:
            fieldnames.append('module_name')

        # 将处理后的数据写入输出CSV文件
        with open(output_file, 'w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            writer.writerows(processed_data)

        print(f"处理成功，结果已保存到 {output_file}")
    except Exception as e:
        print(f"执行命令时出错: {e}")



def merge_ko_and_lsmod(ko_file, lsmod_file, output_file):
    try:
        # 读取 ko_list_sort_modules.csv 文件
        with open(ko_file, 'r', newline='') as ko_file:
            ko_reader = csv.DictReader(ko_file)
            ko_data = list(ko_reader)
            ko_fieldnames = ko_reader.fieldnames

        # 读取 lsmod_output_sort.csv 文件
        with open(lsmod_file, 'r', newline='') as lsmod_file:
            lsmod_reader = csv.DictReader(lsmod_file)
            lsmod_data = list(lsmod_reader)
            lsmod_fieldnames = lsmod_reader.fieldnames

        # 创建一个新的表头，包含 ko_list_sort_modules.csv 的所有列和 lsmod_output_sort.csv 的所有列（去掉重复的 Module Name）
        new_fieldnames = ko_fieldnames + [col for col in lsmod_fieldnames if col != 'Module']

        # 处理每一行数据
        processed_data = []
        for ko_row in ko_data:
            module_name = ko_row['module_name']
            found = False
            for lsmod_row in lsmod_data:
                if lsmod_row['Module'] == module_name:
                    # 合并行数据
                    merged_row = ko_row.copy()
                    for col in lsmod_fieldnames:
                        if col != 'Module':
                            merged_row[col] = lsmod_row[col]
                    processed_data.append(merged_row)
                    found = True
                    break
            if not found:
                # 如果没有找到匹配的行，保留原始行
                processed_data.append(ko_row)

        # 将处理后的数据写入输出CSV文件
        with open(output_file, 'w', newline='') as output_file:
            writer = csv.DictWriter(output_file, fieldnames=new_fieldnames, quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            writer.writerows(processed_data)

        print(f"处理成功，结果已保存到 {output_file.name}")
    except Exception as e:
        print(f"执行命令时出错: {e}")


def process_vmallocinfo_to_csv(output_filename='vmallocinfo.csv'):
    """处理vmallocinfo并输出到CSV文件"""
    # 1. 使用adb shell cat /proc/vmallocinfo 读取内容
    result = subprocess.run(['adb', 'shell', 'cat', '/proc/vmallocinfo'], capture_output=True, text=True)
    lines = result.stdout.splitlines()

    # 2. 如果输出内容小于最长的行数，则缺少的行数默认补齐空格
    padded_lines = []
    for line in lines:
        parts = line.split()

        # 检查第4列是否满足 [*] 格式
        if len(parts) >= 4 and not re.match(r'^\[\w+\]$', parts[3]):
            # 将第4列开始的内容右移动一格，并将第4列填充为空
            parts.insert(3, '')

        # 检查第5列是否包含 pages=* 字符
        if len(parts) >= 5 and not re.match(r'^pages=\d+$', parts[4]):
            # 将第5列开始的内容右移动一格，并将第5列填充为空
            parts.insert(4, '')

        padded_lines.append(parts)

    # 3. 统计输出最长的行，得出最长的行数
    max_length = 0
    for parts in padded_lines:
        length = len(parts)
        if length > max_length:
            max_length = length

    # 4. 再次补齐空格，确保每一行的列数与最长行数一致
    for parts in padded_lines:
        parts.extend([''] * (max_length - len(parts)))

    # 5. 按照第3列的字符数据进行排序
    sorted_lines = sorted(padded_lines, key=lambda x: x[2])

    # 6. 生成表头
    known_headers = ['Address', 'Size', 'Caller', 'Details', 'Pages']
    headers = known_headers + [f'Column_{i + 1}' for i in range(len(known_headers), max_length)]

    # 7. 最终输出到csv文件
    with open(output_filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)  # 写入表头
        writer.writerows(sorted_lines)


def fetch_and_save_slabinfo_to_csv(filename='slabinfo.csv'):
    try:
        # 执行ADB命令获取slabinfo信息
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/slabinfo'], capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")

        # 解析输出
        lines = result.stdout.strip().split('\n')

        # 固定表头
        default_header = [
            'name', 'active_objs', 'num_objs', 'objsize', 'objperslab', 'pagesperslab',
            'tunables', 'limit', 'batchcount', 'sharedfactor', 'slabdata', 'active_slabs',
            'num_slabs', 'sharedavail'
        ]

        # 从第3行开始读取数据
        data = [line.split() for line in lines[2:]]

        # 删除数据中单独的':'符号
        if data:
            # 找到表头中对应的数据列索引
            valid_indices = [i for i, val in enumerate(data[0]) if val != ':']
            data = [[row[i] for i in valid_indices] for row in data]

        # 保存到CSV文件
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(default_header)  # 仅写入表头
            for row in data:
                writer.writerow(row)  # 写入数据

        print(f"Data saved to {filename}")

    except Exception as e:
        print(f"An error occurred: {e}")


def fetch_and_save_iomem_to_csv(filename='iomem.csv'):
    try:
        # 执行ADB命令获取/proc/iomem信息
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/iomem'], capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")

        # 解析输出
        lines = result.stdout.strip().split('\n')

        # 固定表头
        default_header = ['start_address', 'end_address', 'description']

        # 解析数据
        data = []
        for line in lines:
            parts = line.split(':')
            if len(parts) == 2:
                address_range, description = parts
                start_address, end_address = address_range.split('-')
                data.append([start_address.strip(), end_address.strip(), description.strip()])

        # 保存到CSV文件
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL)  # 使用双引号包裹所有字段
            writer.writerow(default_header)  # 写入表头
            for row in data:
                # 确保所有字段以字符串形式写入
                writer.writerow([str(field) for field in row])  # 写入数据

        print(f"Data saved to {filename}")

    except Exception as e:
        print(f"An error occurred: {e}")


def fetch_and_save_vmstat_to_csv(raw_filename='vmstat.txt', csv_filename='vmstat.csv'):
    """
    读取 /proc/vmstat
    1) 保存原始文本到 raw_filename
    2) 转换为 csv_filename (key, value)
    """
    try:
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/vmstat'], capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")

        raw_text = result.stdout or ""
        os.makedirs(os.path.dirname(raw_filename), exist_ok=True)
        with open(raw_filename, mode='w', encoding='utf-8', newline='') as f:
            f.write(raw_text)

        # 解析 vmstat：形如 "nr_free_pages 123456"
        rows = []
        for line in raw_text.splitlines():
            line = line.strip()
            if not line:
                continue
            # 格式：key value 或 key value value2 ...
            parts = line.split()
            if len(parts) >= 2:
                key = parts[0].strip()
                # 取第一个数值作为 value（有些可能有多个值）
                value = parts[1].strip()
                rows.append([key, value])

        os.makedirs(os.path.dirname(csv_filename), exist_ok=True)
        with open(csv_filename, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['key', 'value'])
            for row in rows:
                writer.writerow(row)

        print(f"vmstat raw saved to {raw_filename}")
        print(f"vmstat csv saved to {csv_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")


def fetch_and_save_buddyinfo_to_csv(raw_filename='buddyinfo.txt', csv_filename='buddyinfo.csv'):
    """
    读取 /proc/buddyinfo
    1) 保存原始文本到 raw_filename
    2) 转换为 csv_filename (node, zone, order0, order1, ..., order10)
    """
    try:
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/buddyinfo'], capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")

        raw_text = result.stdout or ""
        os.makedirs(os.path.dirname(raw_filename), exist_ok=True)
        with open(raw_filename, mode='w', encoding='utf-8', newline='') as f:
            f.write(raw_text)

        # 解析 buddyinfo：形如 "Node 0, zone   Normal    123    456    789    ..."
        rows = []
        header = ['node', 'zone', 'order0', 'order1', 'order2', 'order3', 'order4', 'order5', 'order6', 'order7', 'order8', 'order9', 'order10']

        for line in raw_text.splitlines():
            line = line.strip()
            if not line:
                continue

            # 解析格式：Node <num>, zone <name> <order0> <order1> ...
            # 例如：Node 0, zone   Normal    123    456    789    ...
            m = re.match(r'Node\s+(\d+),\s+zone\s+(\S+)\s+(.+)$', line)
            if m:
                node = m.group(1)
                zone = m.group(2)
                values = m.group(3).split()

                # 确保有 11 个值（order0 到 order10）
                while len(values) < 11:
                    values.append('0')

                row = [node, zone] + values[:11]
                rows.append(row)

        os.makedirs(os.path.dirname(csv_filename), exist_ok=True)
        with open(csv_filename, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)
            for row in rows:
                writer.writerow(row)

        print(f"buddyinfo raw saved to {raw_filename}")
        print(f"buddyinfo csv saved to {csv_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")


def fetch_and_save_zoneinfo_to_csv(raw_filename='zoneinfo.txt', csv_filename='zoneinfo.csv'):
    """
    读取 /proc/zoneinfo
    1) 保存原始文本到 raw_filename
    2) 提取关键信息转换为 csv_filename (node, zone, pages_free, pages_min, pages_low, pages_high, ...)
    """
    try:
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/zoneinfo'], capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")

        raw_text = result.stdout or ""
        os.makedirs(os.path.dirname(raw_filename), exist_ok=True)
        with open(raw_filename, mode='w', encoding='utf-8', newline='') as f:
            f.write(raw_text)

        # 解析 zoneinfo：提取每个 zone 的关键信息
        rows = []
        current_node = None
        current_zone = None
        zone_data = {}

        # 定义要提取的关键字段（按优先级排序）
        key_fields = ['pages_free', 'pages_min', 'pages_low', 'pages_high',
                     'pages_spanned', 'pages_present', 'pages_managed',
                     'nr_free_pages', 'nr_zone_inactive_anon', 'nr_zone_active_anon',
                     'nr_zone_inactive_file', 'nr_zone_active_file', 'nr_zone_unevictable',
                     'nr_zone_write_pending', 'nr_mlock', 'nr_page_table_pages',
                     'nr_kernel_stack', 'nr_bounce', 'nr_zspages', 'numa_hit', 'numa_miss']

        def save_current_zone():
            """保存当前 zone 的数据到 rows"""
            nonlocal current_node, current_zone, zone_data, rows
            if current_node is not None and current_zone is not None:
                row = [current_node, current_zone]
                for field in key_fields:
                    row.append(zone_data.get(field, ''))
                rows.append(row)

        for line in raw_text.splitlines():
            line = line.strip()
            if not line:
                continue

            # 解析 Node 和 zone 标题：Node 0, zone   Normal
            m = re.match(r'Node\s+(\d+),\s+zone\s+(\S+)', line)
            if m:
                # 保存上一个 zone 的数据
                save_current_zone()
                # 重置当前 zone 数据
                zone_data = {}
                current_node = m.group(1)
                current_zone = m.group(2)
                continue

            # 解析关键字段：格式可能是 "pages free     123456" 或 "pages_free 123456"
            if current_node is not None and current_zone is not None:
                # 匹配 "key value" 格式，key 可能包含空格或下划线
                # 例如：pages free     123456 或 pages_free 123456
                m2 = re.match(r'^([a-zA-Z_][a-zA-Z0-9_\s]+?)\s+(\d+)$', line)
                if m2:
                    key = m2.group(1).strip().replace(' ', '_')
                    value = m2.group(2).strip()
                    # 只保存我们关心的字段
                    if key in key_fields:
                        zone_data[key] = value

        # 保存最后一个 zone
        save_current_zone()

        # 生成表头
        header = ['node', 'zone'] + key_fields

        os.makedirs(os.path.dirname(csv_filename), exist_ok=True)
        with open(csv_filename, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)
            for row in rows:
                writer.writerow(row)

        print(f"zoneinfo raw saved to {raw_filename}")
        print(f"zoneinfo csv saved to {csv_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")

def fetch_and_save_meminfo(raw_filename='meminfo.txt', csv_filename='meminfo.csv',
                           kernel_filename='meminfo_kernel.csv'):
    """
    读取 /proc/meminfo
    1) 保存原始文本到 raw_filename
    2) 转换为 csv_filename (key,value,unit)
    3) 提取内核空间占用相关项到 kernel_filename，并给出一个估算汇总
    """
    try:
        result = subprocess.run(['adb', 'shell', 'cat', '/proc/meminfo'], capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")

        raw_text = result.stdout or ""
        os.makedirs(os.path.dirname(raw_filename), exist_ok=True)
        with open(raw_filename, mode='w', encoding='utf-8', newline='') as f:
            f.write(raw_text)

        # 解析 meminfo：形如 "MemTotal:       16384256 kB"
        # rows: (key, value, unit) 仅用于落盘到 CSV 时重新组织字段
        rows = []
        meminfo_map = {}
        for line in raw_text.splitlines():
            line = line.strip()
            if not line:
                continue
            m = re.match(r'^([^:]+):\s*([0-9]+)\s*([A-Za-z]+)?\s*$', line)
            if not m:
                continue
            key = m.group(1).strip()
            value = m.group(2).strip()
            unit = (m.group(3) or '').strip()
            rows.append([key, value, unit])
            # 以最新值为准
            try:
                meminfo_map[key] = int(value)
            except Exception:
                pass

        os.makedirs(os.path.dirname(csv_filename), exist_ok=True)
        # 注意：Windows Excel 直接打开 CSV 时，UTF-8 无 BOM 常会显示乱码
        # 这里使用 utf-8-sig 写入 BOM，提升兼容性
        with open(csv_filename, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            # 你要求去掉 unit 列：把单位写在 value 表头里（默认 kb）
            writer.writerow(['key', 'value(kb)'])
            for k, v, _u in rows:
                # meminfo 里绝大多数为 kB；少数无单位项保持原值
                writer.writerow([k, v])

        # =========================
        # 内核空间占用估算（kB）
        #
        # 说明：
        # - /proc/meminfo 不提供“严格精确”的 kernel total（内核/驱动/页缓存/匿名页会有口径差异）
        # - 这里提供一个更贴近“内核空间自身占用”的估算：
        #   Kernel ≈ KReclaimable(若存在) + SUnreclaim + KernelStack + PageTables + Percpu + VmallocUsed + CmaUsed
        # - 其中 KReclaimable 通常比单纯 SReclaimable 更完整（包含更多可回收内核内存）
        # =========================

        memtotal_kb = meminfo_map.get('MemTotal')
        memfree_kb = meminfo_map.get('MemFree')

        # Slab 相关
        slab_kb = meminfo_map.get('Slab')
        s_reclaim_kb = meminfo_map.get('SReclaimable')
        s_unreclaim_kb = meminfo_map.get('SUnreclaim')

        # Slab 缺失时，用 SReclaimable + SUnreclaim 兜底估算
        slab_total_kb = slab_kb
        if slab_total_kb is None and (s_reclaim_kb is not None or s_unreclaim_kb is not None):
            slab_total_kb = (s_reclaim_kb or 0) + (s_unreclaim_kb or 0)

        # 其他典型内核项
        kernelstack_kb = meminfo_map.get('KernelStack')
        pagetables_kb = meminfo_map.get('PageTables')
        percpu_kb = meminfo_map.get('Percpu')
        vmallocused_kb = meminfo_map.get('VmallocUsed')

        # CMA：很多平台会预留/使用较大 CMA，meminfo 中常见
        cma_total_kb = meminfo_map.get('CmaTotal')
        cma_free_kb = meminfo_map.get('CmaFree')
        cma_used_kb = None
        if cma_total_kb is not None and cma_free_kb is not None:
            cma_used_kb = max(cma_total_kb - cma_free_kb, 0)

        # 内核占用估算口径（更稳定）：
        # KernelSpace(内核地址空间分配) ≈ Slab(total) + KernelStack + PageTables + Percpu + VmallocUsed
        # 说明：CMA 是“物理连续内存池/保留区”，常用于 DMA/多媒体缓冲，可能映射到用户态；
        #      是否计入“内核空间占用”因团队口径不同，这里默认【不计入】KernelSpaceEstimatedTotal，
        #      但会额外给出 KernelSpaceEstimatedTotal_WithCMA 作为参考。
        kernel_components_for_sum = [
            ('SlabTotal', slab_total_kb, 'kB', '参与求和；优先 Slab，缺失则用 SReclaimable+SUnreclaim 估算'),
            ('KernelStack', kernelstack_kb, 'kB', ''),
            ('PageTables', pagetables_kb, 'kB', ''),
            ('Percpu', percpu_kb, 'kB', ''),
            ('VmallocUsed', vmallocused_kb, 'kB', '部分平台此项缺失/为0'),
            ('CmaUsed', cma_used_kb, 'kB', 'CmaTotal-CmaFree；默认不计入 KernelSpaceEstimatedTotal'),
        ]
        kernel_sum_present = [(k, v, u, n) for (k, v, u, n) in kernel_components_for_sum if v is not None]
        kernel_total_kb = None
        kernel_total_with_cma_kb = None
        if kernel_sum_present:
            kernel_total_kb = sum(v for k, v, _, _ in kernel_sum_present if k != 'CmaUsed')
            kernel_total_with_cma_kb = sum(v for _, v, _, _ in kernel_sum_present)

        os.makedirs(os.path.dirname(kernel_filename), exist_ok=True)
        with open(kernel_filename, mode='w', newline='', encoding='utf-8-sig') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['key', 'value', 'unit', 'note'])
            # 你要求的顺序：总内存 -> 剩余内存 -> 内核各项 -> 总和/百分比
            if memtotal_kb is not None:
                writer.writerow(['MemTotal', memtotal_kb, 'kB', '总内存'])
            if memfree_kb is not None:
                writer.writerow(['MemFree', memfree_kb, 'kB', '剩余内存（空闲页）'])

            # 内核空间各项
            # 额外输出 Slab 相关原始字段（便于核对）
            if slab_kb is not None:
                writer.writerow(['Slab', slab_kb, 'kB', '原始字段（参考）'])
            if s_reclaim_kb is not None:
                writer.writerow(['SReclaimable', s_reclaim_kb, 'kB', '原始字段（参考）'])
            if s_unreclaim_kb is not None:
                writer.writerow(['SUnreclaim', s_unreclaim_kb, 'kB', '原始字段（参考）'])
            if meminfo_map.get('KReclaimable') is not None:
                writer.writerow(['KReclaimable', meminfo_map.get('KReclaimable'), 'kB', '原始字段（参考）'])

            for k, v, u, n in kernel_sum_present:
                writer.writerow([k, v, u, n])

            if kernel_total_kb is not None:
                writer.writerow(['KernelSpaceEstimatedTotal', kernel_total_kb, 'kB',
                                 '总和=SlabTotal+KernelStack+PageTables+Percpu+VmallocUsed（不含CMA）'])
                if kernel_total_with_cma_kb is not None and (cma_used_kb is not None):
                    writer.writerow(['KernelSpaceEstimatedTotal_WithCMA', kernel_total_with_cma_kb, 'kB',
                                     '参考：含 CmaUsed 的总和（物理连续内存池/保留区）'])
                kernel_total_gb = kernel_total_kb / 1024.0 / 1024.0
                writer.writerow(['KernelSpaceEstimatedTotal_GB', f"{kernel_total_gb:.6f}", 'GB', '1GB=1024*1024kB'])
                if memtotal_kb and memtotal_kb > 0:
                    p = (kernel_total_kb / float(memtotal_kb)) * 100.0
                    writer.writerow(['KernelSpaceEstimatedTotal_PercentOfMemTotal', f"{p:.4f}", '%',
                                     '占比=KernelSpaceEstimatedTotal/MemTotal'])
                else:
                    writer.writerow(['KernelSpaceEstimatedTotal_PercentOfMemTotal', '', '%',
                                     'MemTotal 缺失，无法计算占比'])

                if kernel_total_with_cma_kb is not None and memtotal_kb and memtotal_kb > 0:
                    p2 = (kernel_total_with_cma_kb / float(memtotal_kb)) * 100.0
                    writer.writerow(['KernelSpaceEstimatedTotal_WithCMA_PercentOfMemTotal', f"{p2:.4f}", '%',
                                     '参考占比：KernelSpaceEstimatedTotal_WithCMA/MemTotal'])

        print(f"meminfo raw saved to {raw_filename}")
        print(f"meminfo csv saved to {csv_filename}")
        print(f"meminfo kernel space saved to {kernel_filename}")
    except Exception as e:
        print(f"An error occurred: {e}")


def fetch_and_save_dumpsys_meminfo(raw_filename='dumpsys_meminfo.txt'):
    """
    执行 `adb shell dumpsys meminfo` 获取完整内存信息，并保存到 raw_filename。
    """
    try:
        result = subprocess.run(['adb', 'shell', 'dumpsys', 'meminfo'], capture_output=True, text=True)
        if result.returncode != 0:
            raise Exception(f"Error running command: {result.stderr}")
        raw_text = result.stdout or ""
        os.makedirs(os.path.dirname(raw_filename), exist_ok=True)
        with open(raw_filename, mode='w', encoding='utf-8', newline='') as f:
            f.write(raw_text)
        print(f"dumpsys meminfo raw saved to {raw_filename}")
    except Exception as e:
        print(f"An error occurred: {e}")


def extract_total_pss_by_process_from_dumpsys(raw_filename, output_filename):
    """
    从 dumpsys meminfo 的保存文件中，仅提取 `Total PSS by process` 字段段落保存到 output_filename。
    """
    try:
        if not os.path.isfile(raw_filename):
            raise FileNotFoundError(f"raw dumpsys file not found: {raw_filename}")

        with open(raw_filename, mode='r', encoding='utf-8', errors='replace') as f:
            lines = f.read().splitlines()

        started = False
        captured = []
        for line in lines:
            if not started:
                if re.match(r'^\s*Total\s+PSS\s+by\s+process\s*:\s*$', line, flags=re.IGNORECASE):
                    started = True
                    captured.append(line.rstrip())
                continue

            # started
            # 遇到下一个 "Total PSS by ..." 段落标题则结束
            if re.match(r'^\s*Total\s+PSS\s+by\s+.+:\s*$', line, flags=re.IGNORECASE):
                break
            captured.append(line.rstrip())

        os.makedirs(os.path.dirname(output_filename), exist_ok=True)
        with open(output_filename, mode='w', encoding='utf-8', newline='') as f:
            if captured:
                f.write("\n".join(captured) + "\n")
            else:
                # 找不到段落时也生成文件，避免后续流程报错
                f.write("Total PSS by process:\n")

        print(f"Total PSS by process extracted to {output_filename}")
    except Exception as e:
        print(f"An error occurred: {e}")


def _unit_to_kb_multiplier(unit_letter: str) -> int:
    u = (unit_letter or 'K').upper()
    if u == 'K':
        return 1
    if u == 'M':
        return 1024
    if u == 'G':
        return 1024 * 1024
    return 1


def parse_total_pss_by_process_to_csv(input_filename, output_csv_filename):
    """
    解析 Total PSS by process 段落，形成表格：
      1) 进程名
      2) PID（进程ID）
      3) 内存大小（统一换算为 kB 的数值）
      4) 占总和百分比
      5) 说明（保留原始信息）
    并在最后追加合计行（TOTAL）。
    """
    try:
        if not os.path.isfile(input_filename):
            raise FileNotFoundError(f"input file not found: {input_filename}")

        with open(input_filename, mode='r', encoding='utf-8', errors='replace') as f:
            lines = f.read().splitlines()

        entries = []
        for line in lines:
            if not line.strip():
                continue
            # 跳过标题行
            if re.match(r'^\s*Total\s+PSS\s+by\s+process\s*:\s*$', line, flags=re.IGNORECASE):
                continue

            # 常见格式示例：
            # "  123,456K: com.android.systemui (pid 1234)"
            # "    45,678K: system (pid 567)"
            m = re.match(r'^\s*([0-9][0-9,]*)\s*([KkMmGg])?(?:[Bb])?\s*:\s*(.+?)\s*$', line)
            if not m:
                continue

            num_str = m.group(1)
            unit_letter = m.group(2) or 'K'
            rest = (m.group(3) or '').strip()

            try:
                value = int(num_str.replace(',', ''))
            except Exception:
                continue

            value_kb = value * _unit_to_kb_multiplier(unit_letter)

            # 进程名通常在 "(pid N)" 之前；其余作为说明
            proc_name = rest
            pid = ''
            # 提取 pid 信息，格式可能是 "(pid 1234)" 或 "(pid 1234, ...)"
            pid_match = re.search(r'\(pid\s+(\d+)', rest, flags=re.IGNORECASE)
            if pid_match:
                pid = pid_match.group(1)
                # 从进程名中移除 pid 部分
                pid_split = re.split(r'\s*\(pid\s+\d+.*$', rest, maxsplit=1, flags=re.IGNORECASE)
                if pid_split and pid_split[0].strip():
                    proc_name = pid_split[0].strip()
            else:
                # 如果没有找到 pid，尝试直接使用整个 rest 作为进程名
                proc_name = rest.strip()

            note = rest
            entries.append({
                'process': proc_name,
                'pid': pid,
                'kb': value_kb,
                'note': note,
            })

        total_kb = sum(e['kb'] for e in entries) if entries else 0

        # 输出 CSV（Excel 兼容 UTF-8 BOM）
        os.makedirs(os.path.dirname(output_csv_filename), exist_ok=True)
        with open(output_csv_filename, mode='w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['process_name', 'pid', 'memory_size(kB)', 'percent_of_total', 'note'])

            for e in entries:
                percent = (e['kb'] / float(total_kb) * 100.0) if total_kb > 0 else 0.0
                writer.writerow([e['process'], e['pid'], e['kb'], f"{percent:.4f}%", e['note']])

            # 合计行
            if entries:
                writer.writerow(['TOTAL', '', total_kb, "100.0000%", '合计（Total PSS by process 求和）'])
            else:
                writer.writerow(['TOTAL', '', 0, "", '未解析到有效进程行'])

        print(f"Total PSS by process csv saved to {output_csv_filename}")
    except Exception as e:
        print(f"An error occurred: {e}")


def find_process_path(process_name: str, pid: str) -> str:
    """
    查找进程的原始资源路径（apk 或可执行文件）

    Args:
        process_name: 进程名（可能是包名或可执行文件名）
        pid: 进程ID

    Returns:
        进程路径，如果找不到则返回空字符串
    """
    if not pid:
        return ""

    try:
        # 方法1: 如果是 Android 应用包名（格式如 com.android.xxx），尝试使用 pm path
        if '.' in process_name and not process_name.startswith('/'):
            # 提取包名（去掉组件后缀，如 :provider, :service 等）
            package_name = process_name.split(':')[0] if ':' in process_name else process_name

            result = subprocess.run(
                ['adb', 'shell', 'pm', 'path', package_name],
                capture_output=True, text=True, timeout=10
            )
            if result.returncode == 0 and result.stdout.strip():
                # pm path 输出格式: package:/system_ext/priv-app/OplusLauncher/OplusLauncher.apk
                path = result.stdout.strip().split(':', 1)
                if len(path) == 2:
                    return path[1].strip()

        # 方法2: 通过 /proc/<pid>/exe 查找可执行文件路径
        if pid:
            result = subprocess.run(
                ['adb', 'shell', 'readlink', f'/proc/{pid}/exe'],
                capture_output=True, text=True, timeout=10
            )
            if result.returncode == 0 and result.stdout.strip():
                path = result.stdout.strip()
                if path and path != '':  # readlink 可能返回空或错误
                    return path

            # 方法3: 如果 readlink 失败，尝试 ls -l
            result = subprocess.run(
                ['adb', 'shell', 'ls', '-l', f'/proc/{pid}/exe'],
                capture_output=True, text=True, timeout=10
            )
            if result.returncode == 0 and result.stdout.strip():
                # ls -l 输出格式: lrwxrwxrwx ... -> /system/bin/xxx
                parts = result.stdout.strip().split('->')
                if len(parts) == 2:
                    return parts[1].strip()

        return ""
    except Exception as e:
        # 静默失败，返回空字符串
        return ""


def add_path_column_to_dumpsys_pss(csv_filename: str, max_rows: int = None):
    """
    读取 dumpsys_pss.csv 文件，为每个进程查找路径并添加到 path 列

    Args:
        csv_filename: CSV 文件路径
        max_rows: 最大处理行数（不包括表头），None 表示处理所有行
    """
    try:
        if not os.path.isfile(csv_filename):
            print(f"[WARN] CSV 文件不存在: {csv_filename}")
            return

        # 检查 adb 连接
        result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, timeout=5)
        lines = result.stdout.strip().split('\n')
        devices = [line for line in lines if line.strip() and 'device' in line and 'List' not in line]
        if not devices:
            print("[WARN] 未检测到已连接的设备，跳过路径查找")
            return

        # 读取 CSV 文件
        rows = []
        with open(csv_filename, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            print("[WARN] CSV 文件为空")
            return

        # 查找表头
        header = rows[0]
        if 'path' in header:
            print("[INFO] CSV 文件已包含 path 列，跳过处理")
            return

        # 确定列索引
        try:
            process_name_idx = header.index('process_name')
            pid_idx = header.index('pid')
        except ValueError:
            print("[WARN] CSV 文件缺少必要的列（process_name 或 pid）")
            return

        # 添加 path 列到表头
        header.append('path')

        # 处理数据行
        data_rows = rows[1:]
        if max_rows is not None and max_rows > 0:
            data_rows = data_rows[:max_rows]

        print(f"[INFO] 开始查找进程路径，共 {len(data_rows)} 个进程...")

        for i, row in enumerate(data_rows):
            if len(row) <= max(process_name_idx, pid_idx):
                # 行数据不完整，添加空 path
                row.append('')
                continue

            process_name = row[process_name_idx].strip()
            pid = row[pid_idx].strip()

            # 跳过合计行
            if process_name.upper() == 'TOTAL':
                row.append('')
                continue

            # 查找路径
            path = find_process_path(process_name, pid)
            row.append(path)

            if (i + 1) % 10 == 0:
                print(f"[INFO] 已处理 {i + 1}/{len(data_rows)} 个进程...")

        # 如果限制了行数，保留未处理的行
        if max_rows is not None and max_rows < len(rows) - 1:
            # 保留未处理的行（不添加 path 列，保持原样）
            remaining_rows = rows[max_rows + 1:]
            data_rows.extend(remaining_rows)

        # 写回 CSV 文件
        with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)
            writer.writerows(data_rows)

        print(f"[INFO] 路径查找完成，已更新 {csv_filename}")

    except Exception as e:
        print(f"添加路径列时出错: {e}")


def csv_to_excel(csv_filename, excel_filename=None):
    try:
        # 检查CSV文件是否存在
        if not os.path.isfile(csv_filename):
            raise FileNotFoundError(f"CSV file not found: {csv_filename}")

        # 如果未指定Excel文件名，则使用与CSV文件同名的Excel文件
        if excel_filename is None:
            base_name = os.path.splitext(csv_filename)[0]
            excel_filename = f"{base_name}.xlsx"

        # 读取CSV文件
        df = pd.read_csv(csv_filename, encoding='utf-8-sig')

        # 将数据写入Excel文件（openpyxl 也可能缺失）
        try:
            df.to_excel(excel_filename, index=False, engine='openpyxl')
        except Exception as e:
            print(f"Warning: failed to write excel (openpyxl missing?). Skipping {excel_filename}. ({e})")
            return

        print(f"Data saved to {excel_filename}")

    except Exception as e:
        print(f"An error occurred: {e}")


def _sanitize_excel_sheet_name(name: str) -> str:
    """
    Excel sheet name constraints:
    - Max 31 chars
    - Cannot contain: []:*?/\\
    """
    if name is None:
        name = "Sheet"
    # 替换非法字符
    name = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(name))
    name = name.strip()
    if not name:
        name = "Sheet"
    # 先截断到 31
    return name[:31]


def merge_excel_files_to_one(excel_files, output_excel_filename):
    """
    将多个 Excel 文件按顺序合并到一个 Excel：每个文件一个 Sheet，Sheet 名默认用文件名（不含扩展名）。
    为了便于阅读源码，这里实现为“只拷贝内容”的简化版本：
    - 每个源 xlsx 取第一个 sheet
    - 逐行写入到目标工作簿对应 sheet（按给定顺序）
    - 不复制样式/列宽/冻结窗格/合并单元格等格式
    """
    try:
        wb_out = Workbook()
        # 删除默认 sheet
        wb_out.remove(wb_out.active)

        used_titles = set()

        def _unique_title_from_path(xlsx_path: str) -> str:
            base = os.path.splitext(os.path.basename(xlsx_path))[0]
            base = _sanitize_excel_sheet_name(base)
            if base not in used_titles:
                used_titles.add(base)
                return base
            for i in range(2, 1000):
                suffix = f"_{i}"
                candidate = _sanitize_excel_sheet_name(base[: max(0, 31 - len(suffix))] + suffix)
                if candidate not in used_titles:
                    used_titles.add(candidate)
                    return candidate
            # 兜底
            fallback = _sanitize_excel_sheet_name(base[:29] + "_X")
            used_titles.add(fallback)
            return fallback

        any_added = False
        for xlsx in excel_files:
            if not xlsx:
                continue
            if not os.path.isfile(xlsx) or os.path.getsize(xlsx) <= 0:
                print(f"Warning: Excel not found or empty, skipping: {xlsx}")
                continue

            title = _unique_title_from_path(xlsx)

            try:
                wb_in = load_workbook(xlsx, data_only=False)
                ws_in = wb_in.worksheets[0] if wb_in.worksheets else None
                if ws_in is None:
                    print(f"Warning: empty workbook, skipping: {xlsx}")
                    continue
                ws_out = wb_out.create_sheet(title=title)
                # 逐行拷贝（包括空单元格），以保证列对齐
                for row in ws_in.iter_rows(values_only=True):
                    ws_out.append(list(row))
                any_added = True
                try:
                    wb_in.close()
                except Exception:
                    pass
            except Exception as e:
                print(f"Warning: failed to merge {xlsx}. ({e})")
                continue

        os.makedirs(os.path.dirname(output_excel_filename), exist_ok=True)
        if not any_added:
            # 生成一个空的提示页，避免保存失败
            ws = wb_out.create_sheet(title=_sanitize_excel_sheet_name("EMPTY"))
            ws.cell(row=1, column=1).value = "No excel files merged."

        wb_out.save(output_excel_filename)
        print(f"Merged excel saved to {output_excel_filename}")
    except Exception as e:
        print(f"An error occurred: {e}")


def fetch_and_save_reserved_memory_from_dtsi(output_file: str):
    """
    从手机设备树读取预留内存信息并保存到CSV文件
    参考 fetch_reserved_memory_from_dtsi.py
    """
    try:
        # 检查adb连接
        result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, timeout=5)
        lines = result.stdout.strip().split('\n')
        devices = [line for line in lines if line.strip() and 'device' in line and 'List' not in line]
        if not devices:
            print("[WARN] 未检测到已连接的设备，跳过预留内存信息获取")
            return

        def run_adb_cmd(cmd: str) -> tuple:
            try:
                result = subprocess.run(['adb', 'shell', cmd], capture_output=True, text=True, timeout=30)
                return result.returncode == 0, result.stdout.strip()
            except:
                return False, ""

        def read_reg_hex(node_path: str) -> str:
            success, output = run_adb_cmd(f'cat {node_path}/reg 2>/dev/null | od -An -tx1 -v')
            return output.strip() if success else ""

        def read_size_hex(node_path: str) -> str:
            success, output = run_adb_cmd(f'cat {node_path}/size 2>/dev/null | od -An -tx1 -v')
            return output.strip() if success else ""

        def list_nodes(node_path: str) -> list:
            success, output = run_adb_cmd(f'ls {node_path} 2>/dev/null')
            if success and output:
                return [line.strip() for line in output.split('\n') if line.strip() and not line.startswith('#')]
            return []

        def parse_reg(reg_hex: str) -> tuple:
            if not reg_hex:
                return 0, 0
            hex_bytes = reg_hex.replace(' ', '').strip()
            if len(hex_bytes) < 32:
                return 0, 0
            try:
                addr_high = int(hex_bytes[0:8], 16)
                addr_low = int(hex_bytes[8:16], 16)
                size_high = int(hex_bytes[16:24], 16)
                size_low = int(hex_bytes[24:32], 16)
                addr = (addr_high << 32) | addr_low
                size = (size_high << 32) | size_low
                return addr, size
            except:
                return 0, 0

        def parse_size(size_hex: str) -> int:
            """解析size属性，返回字节数"""
            if not size_hex:
                return 0
            hex_bytes = size_hex.replace(' ', '').strip()
            if len(hex_bytes) < 16:
                return 0
            try:
                # size通常是8字节（64位），大端序
                size_high = int(hex_bytes[0:8], 16)
                size_low = int(hex_bytes[8:16], 16)
                size = (size_high << 32) | size_low
                return size
            except:
                return 0

        def prop_exists(node_path: str, prop: str) -> bool:
            success, output = run_adb_cmd(f'test -f {node_path}/{prop} 2>/dev/null && echo "exists"')
            return success and 'exists' in output

        # 获取所有预留内存节点
        node_names = list_nodes('/proc/device-tree/reserved-memory')
        if not node_names:
            print("[WARN] 无法获取设备树预留内存节点列表")
            return

        nodes = []
        for node_name in node_names:
            node_path = f'/proc/device-tree/reserved-memory/{node_name}'
            # 检查是否是目录
            success, output = run_adb_cmd(f'test -d {node_path} 2>/dev/null && echo "dir"')
            if not (success and 'dir' in output):
                continue

            node_info = {
                'name': node_name,
                'start_addr': 0,
                'end_addr': 0,
                'size_kb': 0,
                'map_type': 'map',
                'reusable': 'non-reusable',
            }

            # 优先读取reg属性（静态定义的地址）
            reg_hex = read_reg_hex(node_path)
            if reg_hex:
                addr, size = parse_reg(reg_hex)
                if addr > 0 and size > 0:
                    node_info['start_addr'] = addr
                    node_info['size_kb'] = size // 1024
                    node_info['end_addr'] = addr + size - 1
            else:
                # 如果没有reg属性，尝试读取size属性（动态分配节点）
                size_hex = read_size_hex(node_path)
                if size_hex:
                    size_bytes = parse_size(size_hex)
                    if size_bytes > 0:
                        node_info['size_kb'] = size_bytes // 1024
                        # 地址保持为0（动态分配，地址在运行时确定）

            # 检查no-map和reusable属性
            if prop_exists(node_path, 'no-map'):
                node_info['map_type'] = 'nomap'
            if prop_exists(node_path, 'reusable'):
                node_info['reusable'] = 'reusable'

            nodes.append(node_info)

        # 过滤：只要有大小就包含（即使地址为0，说明是动态分配）
        valid_nodes = [n for n in nodes if n['size_kb'] > 0]
        # 按地址排序，地址为0的排在后面
        valid_nodes.sort(key=lambda x: (x['start_addr'] if x['start_addr'] > 0 else 0xFFFFFFFFFFFFFFFF, x['name']))

        # 保存到CSV
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['起始地址', '结束地址', '大小(KiB)', '映射类型', '可重用性', '名称'])
            for node in valid_nodes:
                # 如果地址为0，显示为空字符串
                start_addr_str = hex(node['start_addr']) if node['start_addr'] > 0 else ''
                end_addr_str = hex(node['end_addr']) if node['end_addr'] > 0 else ''
                writer.writerow([
                    start_addr_str,
                    end_addr_str,
                    node['size_kb'],
                    node['map_type'],
                    node['reusable'],
                    node['name']
                ])

            # 计算两种统计
            # 1. 静态分配（non-reusable）：不会被释放的内存
            static_nodes = [n for n in valid_nodes if n['reusable'] == 'non-reusable']
            static_total_kb = sum(n['size_kb'] for n in static_nodes)
            static_total_mb = static_total_kb / 1024.0

            # 2. 总分配：所有分配的内存（包括reusable和non-reusable）
            total_kb = sum(n['size_kb'] for n in valid_nodes)
            total_mb = total_kb / 1024.0

            # 添加汇总行
            writer.writerow([])  # 空行分隔
            writer.writerow(['静态分配总计', '', static_total_kb, '', '', '静态分配内存大小(KiB) - non-reusable'])
            writer.writerow(['静态分配总计', '', f"{static_total_mb:.6f}", '', '', '静态分配内存大小(MB) - non-reusable'])
            writer.writerow([])  # 空行分隔
            writer.writerow(['总分配总计', '', total_kb, '', '', '总分配内存大小(KiB) - 所有分配的内存'])
            writer.writerow(['总分配总计', '', f"{total_mb:.6f}", '', '', '总分配内存大小(MB) - 所有分配的内存'])

        print(f"预留内存信息已保存到: {output_file} (有效节点数: {len(valid_nodes)})")
        print(f"[INFO] 静态分配(non-reusable): {static_total_kb} KiB ({static_total_mb:.2f} MB)")
        print(f"[INFO] 总分配(所有内存): {total_kb} KiB ({total_mb:.2f} MB)")
        if len(valid_nodes) < len(nodes):
            skipped = len(nodes) - len(valid_nodes)
            print(f"[INFO] 已跳过 {skipped} 个节点（大小为0）")
    except Exception as e:
        print(f"获取预留内存信息时出错: {e}")


def beautify_excel_file(excel_filename: str, max_width: int = 50, min_width: int = 10, sample_rows: int = 2000):
    """
    对 Excel 做一层“通用可读性”美化：
    - 冻结首行（A2）
    - 首行加粗/底色/居中/自动筛选
    - 列宽自适应（按前 sample_rows 行估算，限制到 [min_width, max_width]）
    - 数字列右对齐 + 千分位
    - TOTAL 行高亮（第一列为 'TOTAL' 时）
    """
    try:
        if not os.path.isfile(excel_filename) or os.path.getsize(excel_filename) <= 0:
            print(f"Warning: excel not found or empty, skipping beautify: {excel_filename}")
            return

        wb = load_workbook(excel_filename)

        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="D9E1F2")  # 淡蓝
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        total_font = Font(bold=True, color="000000")
        total_fill = PatternFill("solid", fgColor="FFF2CC")  # 淡黄

        for ws in wb.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            if max_row <= 0 or max_col <= 0:
                continue

            # 冻结首行（如果有多行）
            if max_row >= 2:
                ws.freeze_panes = "A2"

            # 表头样式（第1行）
            ws.row_dimensions[1].height = 18
            for c in range(1, max_col + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动筛选
            try:
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"
            except Exception:
                pass

            # 判断数字列（基于数据行采样）
            numeric_cols = set()
            scan_rows = min(max_row, 1 + sample_rows)
            for c in range(1, max_col + 1):
                total_cnt = 0
                num_cnt = 0
                for r in range(2, scan_rows + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None or v == "":
                        continue
                    total_cnt += 1
                    if isinstance(v, (int, float)):
                        num_cnt += 1
                if total_cnt > 0 and (num_cnt / float(total_cnt)) >= 0.8:
                    numeric_cols.add(c)

            # 数字列格式（仅对齐，不改变显示格式；避免 3671392 -> 3,671,392）
            num_align = Alignment(horizontal="right", vertical="center")
            txt_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if c in numeric_cols and isinstance(cell.value, (int, float)):
                        cell.alignment = num_align
                    else:
                        # 保持文本为左对齐（避免看起来散）
                        if cell.alignment is None or cell.alignment.horizontal is None:
                            cell.alignment = txt_align

            # TOTAL 行高亮（第一列为 TOTAL）
            try:
                for r in range(2, max_row + 1):
                    v0 = ws.cell(row=r, column=1).value
                    if isinstance(v0, str) and v0.strip().upper() == "TOTAL":
                        for c in range(1, max_col + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.font = total_font
                            cell.fill = total_fill
                        break
            except Exception:
                pass

            # 列宽自适应（按采样行计算）
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                best = 0
                for r in range(1, scan_rows + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    s = str(v)
                    # 经验：中文字符略宽，简单起见按字符数估算
                    best = max(best, len(s))
                # 加一点余量
                width = max(min_width, min(max_width, best + 2))
                ws.column_dimensions[col_letter].width = width

        wb.save(excel_filename)
        print(f"Beautified excel saved to {excel_filename}")
    except Exception as e:
        print(f"Warning: beautify excel failed: {excel_filename}. ({e})")


def main():
    args = parse_cmd_args()
    lsmod_output_sort = os.path.join(args.out, 'lsmod.csv')
    cat_proc_modules_output_sort = os.path.join(args.out, 'modules.csv')
    ko_list_sort_raw = os.path.join(args.out, 'ko_list.csv')
    ko_list_sort_modules = os.path.join(args.out, 'ko_modules.csv')
    merge_ko_and_modules = os.path.join(args.out, 'modules_merge.csv')
    cat_proc_vmallocinfo = os.path.join(args.out, 'vmallocinfo.csv')
    cat_proc_slabinfo = os.path.join(args.out, 'slabinfo.csv')
    cat_proc_iomem = os.path.join(args.out, 'iomem.csv')
    cat_proc_meminfo_raw = os.path.join(args.out, 'meminfo.txt')
    cat_proc_meminfo_csv = os.path.join(args.out, 'meminfo.csv')
    cat_proc_meminfo_kernel = os.path.join(args.out, 'meminfo_kernel.csv')
    dumpsys_meminfo_raw = os.path.join(args.out, 'dumpsys_meminfo.txt')
    dumpsys_total_pss_txt = os.path.join(args.out, 'dumpsys_pss.txt')
    dumpsys_total_pss_csv = os.path.join(args.out, 'dumpsys_pss.csv')
    cat_proc_vmstat_raw = os.path.join(args.out, 'vmstat.txt')
    cat_proc_vmstat_csv = os.path.join(args.out, 'vmstat.csv')
    cat_proc_buddyinfo_raw = os.path.join(args.out, 'buddyinfo.txt')
    cat_proc_buddyinfo_csv = os.path.join(args.out, 'buddyinfo.csv')
    cat_proc_zoneinfo_raw = os.path.join(args.out, 'zoneinfo.txt')
    cat_proc_zoneinfo_csv = os.path.join(args.out, 'zoneinfo.csv')
    reserved_memory_csv = os.path.join(args.out, 'reserved_memory.csv')

    run_adb_lsmod(lsmod_output_sort)
    cat_proc_modules(cat_proc_modules_output_sort)
    find_ko_files(args.list, ko_list_sort_raw)

    # 只有在找到 .ko 文件时才处理
    if os.path.exists(ko_list_sort_raw) and os.path.getsize(ko_list_sort_raw) > 0:
        process_ko_list(ko_list_sort_raw, ko_list_sort_modules)
        merge_ko_and_lsmod(ko_list_sort_modules, lsmod_output_sort, merge_ko_and_modules)
    else:
        print("Warning: No .ko files found, skipping ko_list processing and merge")
        # 创建空的合并文件，避免后续处理出错
        os.makedirs(os.path.dirname(merge_ko_and_modules), exist_ok=True)
        with open(merge_ko_and_modules, 'w', newline='') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['file_name','module_name'])
    process_vmallocinfo_to_csv(cat_proc_vmallocinfo)
    # 示例调用
    fetch_and_save_slabinfo_to_csv(cat_proc_slabinfo)
    fetch_and_save_iomem_to_csv(cat_proc_iomem)
    fetch_and_save_meminfo(cat_proc_meminfo_raw, cat_proc_meminfo_csv, cat_proc_meminfo_kernel)
    fetch_and_save_dumpsys_meminfo(dumpsys_meminfo_raw)
    extract_total_pss_by_process_from_dumpsys(dumpsys_meminfo_raw, dumpsys_total_pss_txt)
    parse_total_pss_by_process_to_csv(dumpsys_total_pss_txt, dumpsys_total_pss_csv)
    # 添加进程路径列
    add_path_column_to_dumpsys_pss(dumpsys_total_pss_csv, max_rows=args.dumpsys_pss_rows)
    fetch_and_save_vmstat_to_csv(cat_proc_vmstat_raw, cat_proc_vmstat_csv)
    fetch_and_save_buddyinfo_to_csv(cat_proc_buddyinfo_raw, cat_proc_buddyinfo_csv)
    fetch_and_save_zoneinfo_to_csv(cat_proc_zoneinfo_raw, cat_proc_zoneinfo_csv)
    fetch_and_save_reserved_memory_from_dtsi(reserved_memory_csv)

    # 所有 CSV 文件列表
    csv_files = [lsmod_output_sort, cat_proc_modules_output_sort, ko_list_sort_raw,
                 ko_list_sort_modules, merge_ko_and_modules, cat_proc_vmallocinfo,
                 cat_proc_slabinfo, cat_proc_iomem, cat_proc_meminfo_csv, cat_proc_meminfo_kernel,
                 dumpsys_total_pss_csv, cat_proc_vmstat_csv, cat_proc_buddyinfo_csv, cat_proc_zoneinfo_csv,
                 reserved_memory_csv]

    for csv_file in csv_files:
        # 只转换存在的 CSV 文件
        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
            csv_to_excel(csv_file)
        else:
            print(f"Warning: CSV file not found or empty: {csv_file}, skipping Excel conversion")

    # =========================
    # 合并多个 Excel（按你给定顺序，合并到不同 Sheet）
    # =========================
    merged_excel = os.path.join(args.out, 'merged_memory_excels.xlsx')
    excel_merge_order = [
        os.path.join(args.out, 'dumpsys_pss.xlsx'),
        os.path.join(args.out, 'modules_merge.xlsx'),
        os.path.join(args.out, 'slabinfo.xlsx'),
        os.path.join(args.out, 'vmallocinfo.xlsx'),
        os.path.join(args.out, 'meminfo.xlsx'),
        os.path.join(args.out, 'meminfo_kernel.xlsx'),
        os.path.join(args.out, 'iomem.xlsx'),
        os.path.join(args.out, 'reserved_memory.xlsx'),
        os.path.join(args.out, 'vmstat.xlsx'),
        os.path.join(args.out, 'buddyinfo.xlsx'),
        os.path.join(args.out, 'zoneinfo.xlsx'),
        os.path.join(args.out, 'modules.xlsx'),
        os.path.join(args.out, 'lsmod.xlsx'),
        os.path.join(args.out, 'ko_modules.xlsx'),
        os.path.join(args.out, 'ko_list.xlsx'),
    ]
    merge_excel_files_to_one(excel_merge_order, merged_excel)
    beautify_excel_file(merged_excel)


if __name__ == "__main__":
    start_time = datetime.now()
    print("Begin time:", start_time.strftime("%Y-%m-%d %H:%M:%S"))
    main()
    end_time = datetime.now()
    print("End time:", end_time.strftime("%Y-%m-%d %H:%M:%S"))
    elapsed_time = end_time - start_time
    print("Total time:", common.format_duration(elapsed_time))


