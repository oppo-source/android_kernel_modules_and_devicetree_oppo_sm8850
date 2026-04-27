# -*- coding: utf-8 -*-
import os
import sys
import requests
import argparse
import hashlib
import time
import random
import hmac
from hashlib import sha1
import mimetypes
from typing import List, Dict, Optional
import platform
from typing import Tuple
from openpyxl import Workbook, load_workbook
from pathlib import Path
import json
import base64
import subprocess
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import re
import ack_json
import api
import pandas as pd
import common

env = "sandbox"
# env = "release"

default_json = {
  "risk_level": "低",
  "module": "默认",
  "analysis": "基本功能",
  "summary": "基本功能",
  "test_recommendations": "基本功能测试"
}
default_header = [
    'revision', 'url', 'message',
    'risk_level', 'module', 'analysis', 'summary', 'test_recommendations'
]

def get_tools_path(relative_path):
    base_path = os.path.dirname(os.path.realpath(__file__))
    full_path = os.path.join(base_path, relative_path)
    if not os.path.exists(full_path):
        os.mkdir(full_path)
    return full_path

def parse_cmd_args():
    parser = argparse.ArgumentParser(description="ACK AI command line tool")
    parser.add_argument('--type', type=str, help='Type of build to send or check (default: sandbox)', default="sandbox")
    parser.add_argument('--method', type=str, help='Method to use for the operation (default: None)', default="")
    parser.add_argument('--host', type=str, help='Host directory for unpacking (default: None)', default="")
    parser.add_argument('--path', type=str, help='Path to the kernel directory (default: None)', default="")
    parser.add_argument('--url_send', type=str, help='URL to send data (default: None)', default="")
    parser.add_argument('--app_id', type=str, help='Application ID (default: None)', default="")
    parser.add_argument('--app_secret', type=str, help='Application secret (default: None)', default="")
    parser.add_argument('--pub', type=str, help='Public key (default: None)', default="")
    parser.add_argument('--secret', type=str, help='Secret key (default: None)', default="")
    parser.add_argument('--message', type=str, help='Message to send (default: "hello world")', default="hello world")
    parser.add_argument('--cmd', type=str, help='Command to execute (default: "send")', default="send")
    parser.add_argument('--project', type=str, help='Project name (default: None)', default="")
    parser.add_argument('--users', type=str, help='Users to notify (default: None)', default="80318998")
    parser.add_argument('--base', type=str, help='base ref', default="android15-6.6-2025-06_r4")
    parser.add_argument('--target', type=str, help='target ref', default="android15-6.6-2025-06_r5")
    parser.add_argument('--gerrit', type=str, help='gerrit dir', default=get_tools_path("../../common"))
    parser.add_argument('-o', '--out', type=str, help='Kernel tmp out dir (default: tools/out)',
                        default=get_tools_path("out"))
    parser.add_argument('--upload_file_id', type=str, help='upload file id',default="")
    args = parser.parse_args()
    """
    for key, value in vars(args).items():
        print(f"{key}: {value}")
    print("")
    """
    return args


def get_api_key(env_type: str, api_name: str) -> str:
    """
    获取指定环境和API名称的认证密钥
    参数:
        env_type: 环境类型 ('sandbox' 或 'release')
        api_name: API名称 (如 'ack_ozone')
    返回:
        对应的Bearer token字符串，未找到时返回空字符串
    """
    env_apis = api.API_KEYS.get(env_type, {})

    key = env_apis.get(api_name, "")

    if not key:
        print("未找到对应的API密钥")

    authorization = f"Bearer {key}"
    return authorization


def get_environment_config(args, env_type, system, path=''):
    config = api.CONFIG.get(env_type, api.CONFIG["default"])
    mtp_paths = api.MTP_PATHS.get("common", api.MTP_PATHS["default"])

    operation_config = config.get(system, config.get("mtp", {}))

    args.host = operation_config.get("host", "default-host")
    args.app_id = operation_config.get("app_id", "default-app-id")
    args.app_secret = operation_config.get("app_secret", "default-app-secret")
    args.secret = operation_config.get("secret", "default-secret")
    args.pub = operation_config.get("pub", "default-pub")
    args.method = "POST"
    args.path = mtp_paths.get(path, 'default-path')
    args.url_send = "https://{}{}".format(args.host, args.path)
    return {
        "mtp": config.get("mtp", {}),
        "ipd-ppm": config.get("ipd-ppm", {})
    }


def hash_hmac(secret, plain_text):
    hmac_code = hmac.new(secret.encode(), plain_text.encode(), sha1).digest()
    return base64.b64encode(hmac_code).decode()

def create_to_content(receive_id, receive_type):
    return {
        "receiveId": receive_id,
        "receiveType": receive_type
    }

def create_request(msg_content, from_content, to_content, msg_type):
    request = {
        "msg": msg_content,
        "from": from_content,
        "to": to_content,
        "type": msg_type
    }
    payload = json.dumps(request, indent=2, ensure_ascii=False)
    #print(payload)
    return payload


def generate_message(user_ids=None, text_content="hello world"):
    # Base message content
    base_message = {
        "tag": "text",
        "text": text_content
    }
    if user_ids and user_ids != []:
        # Dynamically generate @ user parts
        at_users = [
            {
                "tag": "at",
                "userId": user_id
            }
            for user_id in user_ids
        ]
        # Combine message content
        message_content = {
            "content": [
                [base_message] + at_users
            ]
        }
    else:
        # If user_ids is empty or not specified, generate base message content
        message_content = base_message
    return message_content


def modify_headers(headers, action, key=None, value=None):
    """
    通用headers修改函数，支持新增/修改或删除字段

    :param headers: 原始headers字典
    :param action: 操作类型 'add' 或 'remove'
    :param key: 要操作的字段名
    :param value: 要设置的值(仅action='add时需要)
    :return: 修改后的headers字典
    """
    # 创建headers的副本避免修改原始字典
    modified_headers = headers.copy()

    if action == 'add':
        if not key or value is None:
            raise ValueError("key和value参数在action='add时必须提供")
        modified_headers[key] = value  # 存在则更新，不存在则新增

    elif action == 'remove':
        if not key:
            raise ValueError("key参数在action='remove时必须提供")
        modified_headers.pop(key, None)  # 安全删除，不存在也不报错

    else:
        raise ValueError("action参数必须是'add'或remove'")

    return modified_headers

def get_headers_content(args):
    timestamp = int(time.time())
    nonce = str(random.randint(100000, 999999))
    sign_string = "{}&{}&{}&{}&{}&{}".format(args.method, args.host, args.path, args.app_id, timestamp, nonce)
    sign = hash_hmac(args.app_secret, sign_string)
    timestr = str(int(round(time.time() * 1000)))
    list_data = []
    list_data.append(timestr)
    list_data.append(str(nonce))
    list_data.append(str(args.secret))
    list_data.append(str(args.pub))
    list_data.sort()
    str_data = ''.join(list_data)
    pubtoken = hashlib.sha256(str_data.encode('utf-8')).hexdigest()

    str_timestamp = str(timestamp)
    headers = {
        'Content-type': 'application/json',
        'appId': args.app_id,
        'sign': sign,
        'timestamp': str_timestamp,
        'signversion': '2.0.0',
        'host': args.host,
        'nonce': nonce,
    }

    fromcontent = {
        "pub": args.pub,
        "time": timestr,
        "nonce": nonce,
        "pubtoken": pubtoken
    }

    return fromcontent, headers


def create_messages_request(args, users, messages="北京"):
    """
    创建标准化的消息请求JSON
    参数:
        args: 其他参数对象
        users: 用户信息
        messages: 消息内容，默认为"北京"
    返回:
        标准化格式的JSON字符串
    """
    request_data = {
        "query": messages,
        "user": users,
        "inputs": {}
    }

    # 动态添加txt字段
    if hasattr(args, 'message') and args.message:
        request_data["inputs"]["diff_txt"] = args.message

    # 动态添加diff_file字段
    if hasattr(args, 'upload_file_id') and args.upload_file_id:
        request_data["inputs"]["diff_file"] = {
            "type": "document",
            "transfer_method": "local_file",
            "url": "",
            "upload_file_id": args.upload_file_id
        }

    # 如果inputs为空，则移除该字段
    if not request_data["inputs"]:
        request_data.pop("inputs")

    # 转换为标准JSON格式
    json_data = json.dumps(
        request_data,
        indent=2,
        ensure_ascii=False,  # 确保中文正常显示
        sort_keys=True  # 键名排序
    )

    return json_data

def add_authorization_header(headers: dict, key: str = None) -> dict:
    """
    在headers字典中添加Authorization头
    参数:
        headers: 原始headers字典（如果是字符串会自动转换）
        key: 要添加的Authorization值，如果为None则不添加
    返回:
        更新后的headers字典
    """
    # 如果是字符串，转换为字典
    if isinstance(headers, str):
        headers = json.loads(headers)

    # 添加Authorization头
    if key is not None:
        headers['Authorization'] = key

    return headers

def process_excel_hyperlinks(input_path, output_path):
    """
    处理Excel文件：将文本URL转为可点击超链接 + 自动换行 + 智能调整列宽
    参数:
        input_path: 输入Excel文件路径
        output_path: 输出Excel文件路径
    返回:
        bool: 处理是否成功
    """
    # 1. 加载工作簿
    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"❌ 文件加载失败: {str(e)}")
        return False

    # URL匹配正则（增强版）
    url_pattern= re.compile(
        r'https?://(?:[-\w.]|%[\da-fA-F]{2})+(?:[/\w .+-=?&%#]*)?'
    )
    # 2. 处理每个工作表
    for sheet in wb.worksheets:
        print(f"📊 now process data sheet name: {sheet.title}")

        # 3. 处理每个单元格
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue

                # 4. 自动换行设置
                cell.alignment = Alignment(wrapText=True)

                # 5. URL转超链接
                urls = url_pattern.findall(cell.value)
                if urls and cell.value.strip() in urls:
                    cell.hyperlink = urls[0]
                    cell.font = Font(color="0563C1", underline="single")
                    print(f"   🔗 process  url: {urls[0]}")

        # 6. 自动调整列宽
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    # 计算显示长度（中文算2个字符）
                    display_length = sum(2 if ord(c) > 256 else 1 for c in cell_value)
                    if display_length > max_length:
                        max_length = display_length
                except (ValueError, AttributeError) as e:
                    # 只捕获可能发生的特定异常：
                    # - ValueError: str()转换失败
                    # - AttributeError: cell没有value属性
                    print(f"计算单元格显示长度时出错: {e}")
                    max_length = max(max_length, 0)  # 确保max_length不会变小

            # 设置列宽（加2个字符的缓冲）
            adjusted_width = (max_length + 2) * 1.1
            sheet.column_dimensions[col_letter].width = min(adjusted_width, 50)  # 限制最大宽度

    # 7. 保存文件
    try:
        wb.save(output_path)
        print(f"✅ process finished！data save to: {output_path}")
        return True
    except Exception as e:
        print(f"❌ 文件保存失败: {str(e)}")
        return False


def parse_response_answer(result):
    """
    处理JSON数据并提取answer字段

    参数:
        result: JSON格式的字符串

    返回:
        tuple: (标准化后的JSON字符串, answer字段的值)
        如果解析失败则返回 (None, None)
    """
    try:
        # 解析JSON字符串为Python对象
        data_obj = json.loads(result)

        # 提取answer字段
        answer = data_obj.get("answer")
        return answer

    except json.JSONDecodeError as e:
        print(f"JSON解析错误: {e}")
        return None


def parse_response_content(response_data: str, response_type = "message") -> str:
    """
    从SSE(Server-Sent Events)格式的响应中提取有效内容并合并
    参数:
        response_data: 服务器返回的原始响应文本
    返回:
        合并后的有效回答内容字符串
    """
    full_answer = []
    for line in response_data.splitlines():
        if line.startswith('data: '):
            try:
                data = json.loads(line[6:])  # 去掉"data: "前缀
                if data.get('event') == response_type and data.get('answer'):
                    full_answer.append(data['answer'])
            except json.JSONDecodeError:
                continue

    answer = ''.join(filter(None, full_answer))
    print(answer)
    return answer



def parse_event_stream(event_data, event_action = 'workflow_finished'):
    """
    处理工作流事件数据，提取并打印特定信息
    参数:
        event_data: 包含多个事件数据的字符串，每个事件以'event:'开头
    """
    try:
        # 分割事件数据
        events = []
        current_event = {}
        for line in event_data.split('\n'):
            if line.startswith('event:'):
                if current_event:
                    events.append(current_event)
                current_event = {'event': line.split('event:')[1].strip()}
            elif line.startswith('data:'):
                try:
                    data = json.loads(line.split('data:')[1].strip())
                    current_event['data'] = data
                except json.JSONDecodeError as e:
                    print(f"[ERROR] JSON解析错误: {e}")

        if current_event:
            events.append(current_event)

        # 2. 遍历事件并处理
        for event in events:
            event_type = event.get('event')
            if event_type == event_action:
                if event_type == 'workflow_finished':
                    text = event.get('data', {}).get('data', {}).get('outputs', {}).get('text', '')
                    if text:
                        return text
                    else:
                        text = event.get('data', {}).get('data', {}).get('outputs', {}).get('answer', '')
                        if text:
                            return text

                elif event_type == 'agent_thought':
                    text = event.get('data', {}).get('thought', '')
                    if text:
                        return text

    except Exception as e:
        print(f"event: {str(e)}")
        print("错误类型:", type(e).__name__)
        return default_json


def upload_file_to_aisphere(args, headers):
    upload_file = os.path.join(args.out, 'usb.txt')

    if not os.path.exists(upload_file):
        return None

    # 移除Content-type
    headers_without_content_type = modify_headers(
        headers,
        action='remove',
        key='Content-type'
    )

    # 准备文件和表单数据
    with open(upload_file, 'rb') as file:
        # 获取文件的 mimetype
        filename = os.path.basename(upload_file)
        mimetype, _ = mimetypes.guess_type(filename)
        if mimetype is None:
            mimetype = 'application/octet-stream'  # 默认 mimetype

        files = {
            'file': (filename, file, mimetype)
        }
        data = {
            'user': args.users
        }
        response = requests.post(args.url_send, headers=headers_without_content_type, files=files, data=data)

    response_data = json.dumps(response.json(), indent=2, ensure_ascii=False)
    data = json.loads(response_data)
    file_id = data["id"]
    return file_id

def safe_encode(data):
    if isinstance(data, str):
        return data.encode('utf-8')
    elif isinstance(data, dict):
        return {k: v.encode('utf-8') if isinstance(v, str) else v for k, v in data.items()}
    return data



def send_aisphere_request (args, system, api_patch, api_key, event_action="answer", messages="你好"):
    answer = default_json
    try:
        get_environment_config(args, args.type, system, api_patch)
        fromcontent, headers = get_headers_content(args)
        key = get_api_key(args.type, api_key)
        headers = add_authorization_header(headers, key)
        payload = create_messages_request(args, args.users, messages)  # 文本消息
        if api_patch == "upload":
            answer = upload_file_to_aisphere(args,headers)
        else:
            response = requests.post(args.url_send, headers=headers, data=safe_encode(payload))

            data = response.content.decode('utf8')
            if event_action == "answer" or event_action == "completion-blocking-messages":
                answer = parse_response_answer(data)
            elif event_action == "message" or event_action == "agent_message":
                answer = parse_response_content(data, event_action)
            elif event_action == "agent_thought" or event_action == "workflow_finished":
                answer = parse_event_stream(data, event_action)

        return answer

    except Exception as e:
        print(f"event: {str(e)}")
        print("错误类型:", type(e).__name__)
        return answer


def send_tt_normal_message(args, user, msg_type=1, text_content="hello world"):
    if not user or user == '0':
        print("Error: Invalid user. User must be a non-empty string and not '0'.")
        return
    get_environment_config(args, args.type, "mtp", 'send')
    print(args.url_send)
    fromcontent, headers = get_headers_content(args)
    msgcontent = generate_message([], text_content)
    tocontent = create_to_content(user, msg_type)  # 推送给人员
    payload = create_request(msgcontent, fromcontent, tocontent, 2)  # 文本消息
    result = requests.post(args.url_send, headers=headers, data=payload)
    print(result.content.decode('utf8'))

def send_tt_message_cmd(args):
    # 将逗号分隔的字符串转换为列表
    try:
        users_list = args.users.split(',')
    except Exception as e:
        print("Error splitting users: {}".format(e))
        return

    # 去除列表中每个元素的首尾空白字符
    users_list = [user.strip() for user in users_list]

    for user_id in users_list:
        send_tt_normal_message(args, user_id, 1, args.message)


def read_and_print_patch_file(base_path, filename):
    """
    读取并打印单个补丁文件内容

    :param base_path: 基本路径 (如: X:kernel_platform\common)
    :param filename: 补丁文件名
    :return: 文件内容字符串
    """
    # 组合完整路径
    #full_path = os.path.join(base_path, filename)
    full_path = filename

    try:
        # 读取文件内容
        with open(full_path, 'r', encoding='utf-8') as file:
            content = file.read()

        return content

    except FileNotFoundError:
        print(f"警告: 文件未找到 - {full_path}")
        return None
    except Exception as e:
        print(f"读取文件出错 {filename}: {str(e)}")
        return None

def update_excel_file(file_path, new_data, header=None):
    """
    纯Python实现的Excel文件更新（需要openpyxl库）
    """
    if header is None:
        header = default_header
    file_path = Path(file_path)

    if len(new_data) != len(header):
        print(f"[ERROR] 数据列不匹配（需要 {len(header)} 列）")
        return

    try:
        if file_path.exists():
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(header)  # 添加表头

        ws.append(new_data)  # 添加数据
        wb.save(file_path)
        print(f"[SUCCESS] data wrote to: {file_path}")
    except Exception as e:
        print(f"[ERROR] 操作失败: {e}")


def get_commit_diff(
        base_ref: str,
        target_ref: str,
        repo_path: Optional[str] = None,
        include_conflicts: bool = True
) -> List[Dict[str, str]]:
    """
    获取两个分支/标签之间的所有提交差异(包含冲突提交)

    Args:
        base_ref: 基础分支/标签名(旧版本)
        target_ref: 目标分支/标签名(新版本)
        repo_path: Git仓库路径(可选，默认为当前目录)
        include_conflicts: 是否包含冲突解决提交(默认为True)

    Returns:
        包含所有提交信息的列表，每个提交包含revision和message
    """
    # 构建git命令
    git_cmd = ["git"]
    if repo_path:
        git_cmd.extend(["-C", repo_path])

    # 获取两个分支之间的所有提交
    log_format = "'%H %s'"
    cmd = git_cmd + [
        "log",
        f"{base_ref}..{target_ref}",
        "--pretty=format:" + log_format,
        "--no-merges" if not include_conflicts else ""
    ]

    command = " ".join(cmd)
    print(command)
    # 执行命令
    result = subprocess.run(
        command,
        shell=True,
        check=True,
        capture_output=True,
        text=True
    )

    commits = []
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        revision, *message_parts = line.split(' ', 1)
        message = message_parts[0] if message_parts else ""
        commits.append({
            'revision': revision,
            'message': message,
            'is_conflict_resolution': "conflict" in message.lower()  # 简单判断是否为冲突解决
        })

    return commits


def get_os_and_path(windows_path: str, linux_path: str) -> Tuple[str, str]:
    """
    返回当前操作系统类型和对应的路径

    :param windows_path: Windows格式的路径字符串
    :param linux_path: Linux格式的路径字符串
    :return: 元组(操作系统类型, 路径对象)
    :raises FileNotFoundError: 如果路径不存在则抛出异常
    """
    current_os = platform.system().lower()
    os_type = 'windows' if current_os == 'windows' else 'linux'

    path = Path(windows_path) if os_type == 'windows' else Path(linux_path)

    if not path.exists():
        print(f"路径不存在: {path}")

    return os_type, str(path)


def create_named_patch(base_path, commit_hash: str, output_file: str):
    """
    创建指定文件名的补丁

    :param commit_hash: 提交hash（或类似HEAD~1的表达式）
    :param output_file: 输出的补丁文件名（如fix-bug.patch）

    :param base_path: 基础路径
    """
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    cmd = f"git -C {base_path}  format-patch {commit_hash} -1 --stdout > {output_file}"
    subprocess.run(cmd, shell=True, check=True)


# 使用示例
def handle_file_change_from_source(args,ack_changes_list,base_path, filename,revision_id = "000",message=None):

    content = read_and_print_patch_file(base_path, filename)
    if content is not None:
        args.message = content
        answer = send_ack_code_review_chatflow(args)

        if message is None:
            message = content

        url = f"https://android.googlesource.com/kernel/common/+/{revision_id}"
        new_data = [
            revision_id, url, message
        ]
        # 使用示例
        data = ack_json.parse_json_safely(answer)
        # 安全获取字段值（带默认值）
        risk_level = ack_json.handle_json_special_chars(data.get("risk_level", "0"), escape=False)
        module = ack_json.handle_json_special_chars(data.get("module", "0"), escape=False)
        analysis = ack_json.handle_json_special_chars(data.get("analysis", "0"), escape=False)
        summary = ack_json.handle_json_special_chars(data.get("summary", "0"), escape=False)
        test_recommendations = ack_json.handle_json_special_chars(data.get("test_recommendations", "0"), escape=False)

        new_data.extend([risk_level, module, analysis, summary, test_recommendations])
        update_excel_file(ack_changes_list, new_data, default_header)


# 使用示例
def analysis_ack_changes(args):
    ack_changes_list_init = os.path.join(args.out, 'ack_changes_list_init.xlsx')
    ack_changes_list_init_hyperlinks = os.path.join(args.out, 'ack_changes_list_init_hyperlinks.xlsx')
    ack_changes_list_merged = os.path.join(args.out, 'ack_changes_list_merged.xlsx')
    ack_changes_list_merged_hyperlinks = os.path.join(args.out, 'ack_changes_list_merged_hyperlinks.xlsx')
    ack_changes_list_merged_consolidate = os.path.join(args.out, 'ack_changes_list_merged_consolidate.xlsx')
    ack_changes_list_merged_consolidate_hyperlinks = os.path.join(args.out, 'ack_changes_list_merged_consolidate_hyperlinks.xlsx')

    merge_excel_data(ack_changes_list_init, ack_changes_list_merged)
    process_excel_hyperlinks(ack_changes_list_init, ack_changes_list_init_hyperlinks)

    analysis_final_data(args, ack_changes_list_merged,ack_changes_list_merged_consolidate)

    process_excel_hyperlinks(ack_changes_list_merged, ack_changes_list_merged_hyperlinks)
    process_excel_hyperlinks(ack_changes_list_merged_consolidate, ack_changes_list_merged_consolidate_hyperlinks)

# 使用示例
def read_changes_from_source(args):
    os_type = ""
    project_path = ""
    ack_changes_list_init = os.path.join(args.out, 'ack_changes_list_init.xlsx')
    if os.path.exists(ack_changes_list_init):
        os.remove(ack_changes_list_init)

    # 使用示例
    try:
        os_type, project_path = get_os_and_path(
            windows_path=r"Z:\code\8850\source\vnd\kernel_platform\common",
            linux_path = args.gerrit
        )
        print(f"os: {os_type}, path: {project_path}")
    except FileNotFoundError as e:
        print(f"错误: {e}")

    base_path = str(project_path)

    if not os.path.exists(base_path):
        print(f"path: {base_path} not exists")
        return

    if os_type == 'linux':
        # 带路径和包含冲突提交的用法
        diff = get_commit_diff(
            base_ref=args.base,
            target_ref=args.target,
            repo_path=base_path,
            include_conflicts=True
        )
        total_commits = len(diff)
        print(f"total: {total_commits} commit")
        for index, commit in enumerate(diff, start=1):
            print(f"now process {index}/{total_commits} commit")
            revision = commit['revision']
            message = commit['message']
            print(f"revision: {revision} message: {message}")
            output= os.path.join(base_path, 'diff_revision.patch')
            create_named_patch(
                base_path,
                commit_hash=revision,
                output_file=output
            )
            try:
                handle_file_change_from_source(args, ack_changes_list_init, base_path, output,revision, message)
            except Exception as e:
                print(f"Error: {e}")
            print("-" * 50)

    else:
        patch_files = [
            "0001-usb-dwc3-gadget-Rewrite-endpoint-allocation-flow.patch"
        ]
        for filename in patch_files:
            print(f"filename: {filename}")
            handle_file_change_from_source(args, ack_changes_list_init, base_path, filename)

    analysis_ack_changes(args)

def merge_excel_data(input_file, output_file):
    """
    使用标准库处理Excel文件，合并相同module的summary和test_recommendations

    参数:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径
    """
    # 加载工作簿
    wb = load_workbook(input_file)
    ws = wb.active

    # 获取列索引
    headers = [cell.value for cell in ws[1]]
    module_idx = headers.index('module')
    summary_idx = headers.index('summary')
    test_idx = headers.index('test_recommendations')

    # 创建合并数据字典
    merged_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        module = row[module_idx]
        if module not in merged_data:
            merged_data[module] = {
                'summary': [],
                'test_recommendations': []
            }
        merged_data[module]['summary'].append(str(row[summary_idx]))
        merged_data[module]['test_recommendations'].append(str(row[test_idx]))

    # 创建新工作簿
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['module', 'summary', 'test_recommendations'])

    # 写入合并后的数据
    for module, data in merged_data.items():
        new_ws.append([
            module,
            '\n'.join(data['summary']),
            '\n'.join(data['test_recommendations'])
        ])

    # 保存新文件
    new_wb.save(output_file)


def analysis_final_data(args, input_file, output_file=None):
    """
    读取Excel文件并打印module、summary和test_recommendations列的内容，可选输出到新Excel文件
    参数:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径(可选)
    """
    wb = load_workbook(input_file)
    ws = wb.active

    # 初始化输出工作簿和工作表变量
    output_wb = None
    output_ws = None

    # 如果需要输出到文件，创建新工作簿
    if output_file:
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.append(['Module', 'Summary', 'Test Recommendations'])

    # 获取列索引
    headers = [cell.value for cell in ws[1]]
    module_idx = headers.index('module')
    summary_idx = headers.index('summary')
    test_idx = headers.index('test_recommendations')

    # 计算总行数
    total_rows = ws.max_row - 1  # 减去表头行
    print(f"Total rows to process: {total_rows}")
    print("-" * 85)

    # 逐行处理内容
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        module = str(row[module_idx]) if row[module_idx] is not None else ""
        summary = str(row[summary_idx]) if row[summary_idx] is not None else ""
        test_rec = str(row[test_idx]) if row[test_idx] is not None else ""

        print(f"Processing row summary {row_num-1}/{total_rows}")  # +1因为从第2行开始
        args.message = summary
        summary = send_ack_review_summary_analysis_chatflow(args)

        print(f"Processing row test recommend {row_num-1}/{total_rows}")  # +1因为从第2行开始
        args.message = test_rec
        test_rec = send_ack_review_test_case_analysis_chatflow(args)

        if output_file and output_ws:
            output_ws.append([module, summary, test_rec])

    # 保存输出文件
    if output_file and output_wb:
        output_wb.save(output_file)


def send_search_capital_assistant_messages (args):
    answer = send_aisphere_request(args, "mtp", 'chat-blocking-messages',
                          "search_capital_assistant",  messages="北京")
    print(answer)

def send_search_ack_local_knowledge (args):
    answer = send_aisphere_request(args, "mtp", 'chat-blocking-messages',
                          "ack_local_knowledge", messages="OKI")
    print(answer)

def send_chat_assistant (args):
    answer = send_aisphere_request(args, "mtp", 'chat-blocking-messages',
                          "chat_assistant", messages="你好")
    print(answer)

def send_search_ack_ozone (args):
    answer = send_aisphere_request(args, "mtp", 'chat-blocking-messages',
                          "ack_ozone", messages="编译")
    print(answer)

def send_ack_code_review_workflow (args):
    answer = send_aisphere_request(args, "mtp", 'streaming-messages',
                          "ack_code_review_workflow", event_action = "workflow_finished", messages="你好")
    print(answer)

def send_ack_code_review_chatflow (args):
    answer= send_aisphere_request(args, "mtp", 'streaming-messages',
                          "ack_code_review_chatflow", event_action = "workflow_finished",
                          messages="请review this code,首先判断风险等级")
    return answer

def send_ack_review_test_case_analysis_chatflow (args):
    answer= send_aisphere_request(args, "mtp", 'streaming-messages',
                          "ack_review_test_case_analysis_chatflow", event_action = "workflow_finished",
                          messages="请分析这个数据集，并给出分析结果")
    return answer

def send_ack_review_summary_analysis_chatflow (args):
    answer= send_aisphere_request(args, "mtp", 'streaming-messages',
                          "ack_review_summary_analysis_chatflow", event_action = "workflow_finished",
                          messages="请分析这个总结，并给出分析结果")
    return answer

def send_ack_patch_upload (args):
    answer = send_aisphere_request(args, "mtp", 'upload',
                          "ack_code_review_chatflow", event_action = "workflow_finished", messages="你好")
    if answer is not None:
        print("ack_patch_upload uuid:" + answer)
        args.upload_file_id = answer
        answer = send_ack_code_review_chatflow(args)
        print(answer)
    else:
        print("answer is None ignore")

def send_code_generate_assistant (args):
    answer = send_aisphere_request(args, "mtp", 'completion-blocking-messages',
                          "code_generate_assistant", messages="请输出python参考示例")
    print(answer)

def send_agent_demo_assistant (args):
    answer = send_aisphere_request(args, "mtp", 'streaming-messages',
                          "agent_demo_assistant", event_action = "agent_thought", messages="你好")
    print(answer)

def git_pull_sourcetree(args):
    command = ['git', '-C', args.gerrit, 'pull']
    try:
        result = subprocess.run(command,
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                universal_newlines=True,
                                check=True)
    except subprocess.CalledProcessError as e:
        err = e.stderr.strip()
        path = repo_path or os.getcwd()
        if "not a git repository" in err.lower():
            raise subprocess.SubprocessError("非Git仓库: {}".format(path)) from e
        raise subprocess.SubprocessError("Git错误: {}".format(err)) from e

def merge_excel_files(file1, file2, output_file):
    """
    将两个Excel文件合并成一个Excel文件的不同工作表

    参数:
    file1: 第一个Excel文件路径
    file2: 第二个Excel文件路径
    output_file: 输出文件路径
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file1):
            print(f"错误: 文件 {file1} 不存在")
            return
        if not os.path.exists(file2):
            print(f"错误: 文件 {file2} 不存在")
            return

        # 创建Excel写入对象
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

            # 读取并写入第一个文件
            print(f"正在处理文件: {file1}")
            df1 = pd.read_excel(file1)
            # 使用文件名（不含扩展名）作为工作表名
            sheet_name1 = os.path.splitext(os.path.basename(file1))[0]
            sheet_name1 = re.sub(r'^ack_changes_list_', '', sheet_name1)
            # 限制工作表名长度（Excel限制为31个字符）
            sheet_name1 = sheet_name1[:31]
            df1.to_excel(writer, sheet_name=sheet_name1, index=False)
            print(f"已添加工作表: {sheet_name1}，包含 {len(df1)} 行数据")

            # 读取并写入第二个文件
            print(f"正在处理文件: {file2}")
            df2 = pd.read_excel(file2)
            sheet_name2 = os.path.splitext(os.path.basename(file2))[0]
            sheet_name2 = re.sub(r'^ack_changes_list_', '', sheet_name2)
            sheet_name2 = sheet_name2[:31]
            df2.to_excel(writer, sheet_name=sheet_name2, index=False)
            print(f"已添加工作表: {sheet_name2}，包含 {len(df2)} 行数据")

        print(f"\n合并完成！输出文件: {output_file}")

    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")

def update_files(args):
    jfrog_url = "xxx/aosp-gki-image-local/ack_ai_analysis"
    merged_hyperlinks = os.path.join(args.out, "ack_changes_list_init_hyperlinks.xlsx")
    merged_consolidate_hyperlinks = os.path.join(args.out, "ack_changes_list_merged_consolidate_hyperlinks.xlsx")

    merge_file = "ack_ai_analysis_" + args.base + "_" + args.target + ".xlsx"
    local_file = os.path.join(args.out, merge_file)
    merge_excel_files(merged_hyperlinks, merged_consolidate_hyperlinks, local_file)

    remote_file = os.path.join(jfrog_url, merge_file)
    common.upload_file(local_file, remote_file)

def main():
    args = parse_cmd_args()
    git_pull_sourcetree(args)
    #send_search_capital_assistant_messages(args)
    #send_search_ack_local_knowledge(args)
    #send_chat_assistant(args)
    #send_search_ack_ozone(args)
    #send_ack_code_review_workflow(args)
    #send_ack_patch_upload(args)
    #send_code_generate_assistant(args)
    #send_agent_demo_assistant(args)
    #send_tt_message_cmd(args)
    read_changes_from_source(args)
    #analysis_ack_changes(args)
    #update_files(args)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
